#!/usr/bin/env python3

# PyAssess2026.py - generates processed exam grids for Physics@Manchester
# Author: Clive Dickinson
# Date: 2026-05-30
# Version: 1.0 (04-Jun-2026 almost ready for June exams - resits will do over the summer)

# Requirements: 
# -Python >3.10
# -pandas v2.1.0 or higher
# -openpyxl v3.1.2 or higher
#
# How to run:
# Python PyAssess2026.py --classyear 1/1m/2m/31/31m/4/4m/all [ --fill_marks --sort ]
# (change AY, INDIR and FILENAMES as appropriate)

import argparse
import math
import os
import re
import statistics
import sys

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ===========================================================================
# Constants — update as required
# ===========================================================================
AY           = 2026   # default academic year (2025 = AY 2024-25 etc.); override on the command line with --AY
OUTDIR       = "./"   # Output directory for generated spreadsheets
SORT_OUTPUT  = False  # Sort output by mark (descending); overridden by --sort
FILL_MARKS   = None   # Fill blank marks before processing: None = disabled, or a float e.g. 50.0; overridden by --fill_marks


def _configure_ay(ay):
    """Set every academic-year-dependent global from *ay*.

    Centralises all per-year configuration: the input directory, the
    supplementary filenames, the special-status student lists, and the
    classyear -> (input_file, output_file) map.  Called once at import with the
    module default (AY), and again from main() when --AY overrides it.  Exits on
    an unsupported year.

    Special-status lists hold emplids (int or str) that override normal
    processing: matched students have their status (progressing years) or Deg
    Class Alg/Actual (final years) set to the label, computed output codes
    cleared, and resits/fail reasons removed.  abroad_list holds 4-year
    MPhys/MMath students who spent a year abroad but whose plan string does not
    contain 'study' (so they aren't auto-detected).
    """
    global AY, INDIR, Y3_CREDITS_FILE, CF_FLAG_FILE, ABROAD_FILE
    global interrupt_list, manual_list, withdrawn_list, abroad_list
    global CLASSYEAR_FILES, ALL_CLASSYEARS

    if ay == 2025:
        INDIR           = "data2025_newcodes"
        Y3_CREDITS_FILE = "y4-comp-and-L4-2025.xlsx"   # supplementary Y3 credit data for Y4 students
        CF_FLAG_FILE    = 'PHYS Carry forward.xlsx'    # carry-forward notes (None to disable)
        ABROAD_FILE     = 'Study abroad 2023-24.xlsx'  # extra study-abroad emplids (None to disable)
        interrupt_list  = [11085845, 11021684, 11061389, 11064835]   # -> 'Interrupt'
        manual_list     = []                                         # -> 'Manual'
        withdrawn_list  = [9819578, 11307037, 10895432]              # -> 'Withdrawn'
        abroad_list     = [10820251, 10826844, 10830728, 10849791, 10818157,
                           10821713, 10833591, 10834291, 10943429, 10669557,
                           10706596, 11487593, 11498024, 11498021, 11498023]
        CLASSYEAR_FILES = {
            '1':   ('PHYS_1241_S2_Y1_Exam_Grids.xlsx',         f'1styear_Physics.AY{ay}.xlsx'),
            '1m':  ('PHYS_1241_S2_Y1_MP_Exam_Grids.xlsx',      f'1styear_MathsPhysics.AY{ay}.xlsx'),
            '2':   ('PHYS_1241_S2_Y2_Exam_Grids.xlsx',         f'2ndyear_Physics.AY{ay}.xlsx'),
            '2m':  ('PHYS_1241_S2_Y2_MP_Exam_Grids.xlsx',      f'2ndyear_MathsPhysics.AY{ay}.xlsx'),
            '31':  ('PHYS_1241_S2_Y3_PROG_Exam_Grids.xlsx',    f'3rdyear_MPhys.AY{ay}.xlsx'),
            '31m': ('PHYS_1241_S2_Y3_MP_PROG_Exam_Grids.xlsx', f'3rdyear_MMath.AY{ay}.xlsx'),
            '32':  ('PHYS_1241_S2_Y3_GRAD_Exam_Grids.xlsx',    f'FinalYear_BSc_Physics.AY{ay}.xlsx'),
            '32m': ('PHYS_1241_S2_Y3_MP_GRAD_Exam_Grids.xlsx', f'FinalYear_BSc_MathsPhysics.AY{ay}.xlsx'),
            '4':   ('PHYS_1241_S2_Y4_Exam_Grids.xlsx',         f'FinalYear_MPhys.AY{ay}.xlsx'),
            '4m':  ('PHYS_1241_S2_Y4_MP_Exam_Grids.xlsx',      f'FinalYear_MMath.AY{ay}.xlsx'),
        }
    elif ay == 2026:
        INDIR           = "data2026"
        Y3_CREDITS_FILE = "level4&comp-2026.xlsx"
        CF_FLAG_FILE    = 'A6_A10.xlsx'
        ABROAD_FILE     = '1251_PHYS_Study_Abroad.xlsx'
        interrupt_list  = []
        manual_list     = []
        withdrawn_list  = []
        abroad_list     = []
        CLASSYEAR_FILES = {
            '1':   ('Y1.xlsx',         f'1styear_Physics.AY{ay}.xlsx'),
            '1m':  ('Y1_MP.xlsx',      f'1styear_MathsPhysics.AY{ay}.xlsx'),
            '2':   ('Y2.xlsx',         f'2ndyear_Physics.AY{ay}.xlsx'),
            '2m':  ('Y2_MP.xlsx',      f'2ndyear_MathsPhysics.AY{ay}.xlsx'),
            '31':  ('PHYS_1251_S2_Y3_PROG_EXAM_GRID.xlsx',    f'3rdyear_MPhys.AY{ay}.xlsx'),
            '31m': ('PHYS_1251_S2_Y3_MP_PROG_EXAM_GRID.xlsx',  f'3rdyear_MMath.AY{ay}.xlsx'),
            '32':  ('PHYS_1251_S2_Y3_GRAD_EXAM_GRID.xlsx',    f'FinalYear_BSc_Physics.AY{ay}.xlsx'),
            '32m': ('PHYS_1251_S2_Y3_MP_GRAD_EXAM_GRID.xlsx',  f'FinalYear_BSc_MathsPhysics.AY{ay}.xlsx'),
            '4':   ('PHYS_1251_S2_Y4_GRAD_EXAM_GRID.xlsx',         f'FinalYear_MPhys.AY{ay}.xlsx'),
            '4m':  ('PHYS_1251_S2_Y4_MP_GRAD_EXAM_GRID.xlsx',       f'FinalYear_MMath.AY{ay}.xlsx'),
        }
    else:
        sys.exit(f"Error: unsupported --AY {ay}; valid academic years: 2025, 2026")

    AY = ay
    ALL_CLASSYEARS = list(CLASSYEAR_FILES.keys())


_configure_ay(AY)   # populate AY-dependent globals at import; main() re-runs it if --AY is given

# Pass mark for any individual unit
PASS_MARK = 39.95

# Minimum mark in any unit to avoid outright fail (30%)
MIN_MARK = 29.95

# Degree classification boundaries
BOUNDARY_FIRST  = 69.95
BOUNDARY_UPPER2 = 59.95
BOUNDARY_LOWER2 = 49.95
BOUNDARY_THIRD  = 39.95

# BSc L3 credit requirements for degree classification.
# For 1st/2.1/2.2 the student must have passed >= BSC_L3_CREDITS_UPPER credits at
# level 3 (mark > PASS_MARK), including all MUST_PASS units (lab and project).
# For a proper 3rd they need >= BSC_L3_CREDITS_THIRD L3 credits including MUST_PASS units;
# for an ordinary 3rd they need >= BSC_L3_CREDITS_THIRD without the MUST_PASS requirement.
BSC_L3_CREDITS_UPPER = 80
BSC_L3_CREDITS_THIRD = 60

# Borderline zone width below each upper-class boundary (2%) and the third boundary (3%).
# A student whose overall mark falls within this band below the next boundary is considered
# borderline and may be promoted by algorithms A or B.
BSC_BORDERLINE_UPPER = 2.0   # applies to 1st, 2.1, 2.2 boundaries
BSC_BORDERLINE_THIRD = 3.0   # applies to the 3rd-class boundary

# Minimum credits (any current-year level, at marks >= the target boundary) for
# borderline promotion.  Both programmes try Algorithm A first, then Algorithm B.
# Algorithm A's threshold differs by programme (BSC_PROMO_A_CREDITS for BSc,
# Y4_PROMO_A_CREDITS for MPhys/MMath); Algorithm B shares one threshold
# (PROMO_B_CREDITS) but also needs the project at the target level and yearmark > overall.
BSC_PROMO_A_CREDITS = 80   # BSc algorithm A
Y4_PROMO_A_CREDITS  = 75   # MPhys/MMath algorithm A
PROMO_B_CREDITS     = 70   # algorithm B (BSc and MPhys/MMath)

# MPhys/MMath (4-year) degree classification credit requirements.
# Classification mirrors the BSc rules but over the combined Y3+Y4 (level 3 and
# above) credits, assuming 240 credits taken across the two years.  A student must
# pass >= MPHYS_CREDITS_FULL credits (including the project) for the class their
# overall mark implies; if short by up to 20 credits (>= MPHYS_CREDITS_SHORT) and
# the project is passed, the class is dropped one level.  There is no 3rd class:
# anyone who fails the 2nd-class requirements, or scores below the 2.2 boundary
# without being promoted, reverts to a BSc based on their first three years.
MPHYS_CREDITS_TOTAL = 240    # nominal Y3+Y4 credits taken
MPHYS_CREDITS_FULL  = 200    # credits passed (incl. project) for the mark-indicated class
MPHYS_CREDITS_SHORT = 180    # short-credit floor (up to 20 short) → one class lower

# BSc dissertation modules — used to locate the dissertation unit for Algorithm B.
BSC_DISSERTATION_MODULES = frozenset({'PHYS30880', 'PHYS30881', 'PHYS30882'})

# BSc project modules: the Physics dissertation, plus MATH30022 — the equivalent
# project for BSc Maths+Physics students.  These are the BSc project mark.
BSC_PROJECT_MODULES = BSC_DISSERTATION_MODULES | frozenset({'MATH30022'})

# Professional placement and year-abroad module codes.
# If a student has one of these units and the mark is absent (blank, -1, or
# a non-numeric string such as 'P'/'F'), they are classified as 'Intercal'.
PP_MODULES     = frozenset({'PHYS30810', 'PHYS40810'})
ABROAD_MODULES = frozenset({'PHYS31000', 'PHYS41000'})

# Credit thresholds for Y1/Y2 progression
MIN_CREDITS_TO_PROGRESS  = 80   # credits at >= PASS_MARK needed to progress without resits
MIN_PASS_CREDITS         = 60   # credits at >= PASS_MARK needed at first attempt to avoid FAIL

# Y2 yearmark thresholds for MPhys/MMath progression
MPHYS_PROGRESS_MARK = 54.95   # must exceed this to progress to MPhys/MMath Y3 (ACTV)
MPHYS_REVIEW_MARK   = 52.95   # borderline band: >this but <=MPHYS_PROGRESS_MARK → R/X
                               # at or below this → BSc (moved to BSc programme)

# Y3 MPhys/MMath → Y4 progression thresholds (classyear 31/31m)
MPHYS_Y3_PROG_YEARMARK       = 49.95   # yearmark must exceed this to progress
MPHYS_Y3_PROG_OVERALL        = 49.95   # overall mark must exceed this to progress
MPHYS_Y3_PROG_CREDITS        = 100     # minimum credits passed (mark > PASS_MARK) to progress
MPHYS_Y3_PROG_PHYS_MATH_MARK = 44.95  # MMath only: both phys and maths yearmarks must exceed this

FINAL_CLASSYEARS = ['32', '32m', '4', '4m']   # graduating / final-year students

# Lab units that must be passed if taken — the experimental physics lab in each
# year/semester.  A failed lab (in any year) cannot be compensated.
MUST_PASS_LAB = frozenset({'PHYS10180', 'PHYS10280', 'PHYS20180', 'PHYS20280',
                           'PHYS30180', 'PHYS30280'})

# Project units that must be passed for the degree — the BSc project
# (BSC_PROJECT_MODULES) and the MPhys project (PHYS40181/2).  These form the
# project mark and are a pass requirement for the BSc and MPhys/MMath degrees.
MUST_PASS_PROJECT = BSC_PROJECT_MODULES | frozenset({'PHYS40181', 'PHYS40182'})

# All units that must be passed if taken (lab in any year + project).
MUST_PASS = MUST_PASS_LAB | MUST_PASS_PROJECT

# Project / dissertation modules for OUTPUT ordering — the must-pass projects plus
# the alternative project units that fill a project slot (PHIL40000 for Physics with
# Philosophy, MATH40011/MATH40022 for MMath). Used by _order_units to place project
# units after the lab in the output grid. (See calc_project_mark for the same codes.)
PROJECT_MODULES = MUST_PASS_PROJECT | frozenset({'PHIL40000', 'MATH40011', 'MATH40022'})

# Maths units that Y1 M+P students must pass — cannot be compensated
MUST_PASS_MATHS = frozenset({'MATH11121', 'MATH11022'})

# Core Physics units (required for resit decisions)
CORE_PHYSICS = frozenset({
    'PHYS10071',
    'PHYS10101',
    'PHYS10121',
    'PHYS10191',
    'PHYS10302',
    'PHYS10342',
    'PHYS10372',
    'PHYS20101',
    'PHYS20141',
    'PHYS20151',   # Y2 changed in 2025 (replaces PHYS20161 - Judith email 30-Jun-2025)
    'PHYS20171',
    'PHYS20302',   # Y2 changed in 2025 (replaces PHYS20252 - Judith email 06-May-2025)
    'PHYS20342',   # Y2 changed in 2025 (replaces PHYS20312 - Judith email 06-May-2025)
    'PHYS20352',
})

# Additional core maths units for Maths+Physics students
CORE_MATHS_PHYSICS = frozenset({
    'MATH10111', 'MATH10121', 'MATH10212', 'MATH11222',
    'MATH11121', 'MATH11022', 'MATH29141', 'MATH24420',
})

# ===========================================================================
# Per-class-year file mappings
# ===========================================================================
# classyear keys:
#   1 / 1m  = Year 1  Physics / Maths+Physics
#   2 / 2m  = Year 2  Physics / Maths+Physics
#  31 / 31m = Year 3  MPhys progressing / MMath progressing
#  32 / 32m = Year 3  BSc Physics graduating / BSc MathsPhysics graduating
#   4 / 4m  = Year 4  MPhys (final year) / MMath (final year)
#
# Values: (input_filename, output_filename)

# CLASSYEAR_FILES (classyear -> (input_file, output_file)) and ALL_CLASSYEARS are
# populated per academic year by _configure_ay() near the top of this file.

# Human-readable description for each classyear key (used in reports).
_CY_DESC = {
    '1':   'Y1 Physics',
    '1m':  'Y1 Maths+Physics',
    '2':   'Y2 Physics',
    '2m':  'Y2 Maths+Physics',
    '31':  'Y3 MPhys (progressing)',
    '31m': 'Y3 MMath (progressing)',
    '32':  'Y3 BSc Physics',
    '32m': 'Y3 BSc Maths+Physics',
    '4':   'Y4 MPhys',
    '4m':  'Y4 MMath',
}

# ===========================================================================
# Excel sheet layout (same across all input files)
# ===========================================================================
SHEET_NAME     = 'Style A Plus With Gradebook'
UNIT_LABEL_ROW = 4   # 0-indexed: row 5 — contains 'Unit 1', 'Unit 2', ...
COL_HEADER_ROW = 5   # 0-indexed: row 6 — contains 'Module', 'Mark', 'EN', ...
DATA_START_ROW = 6   # 0-indexed: row 7 — first student row

# Within each unit block the four columns of interest sit at these offsets
# from the unit's start column (Module, Link1, Link2, Mark, EN, Mit Circs, ...)
_UNIT_COL_OFFSETS = {'module': 0, 'mark': 3, 'en': 4, 'mit_circs': 5}

# ===========================================================================
# Data classes
# ===========================================================================

_MODULE_RE      = re.compile(r'^(.*?)\s*\((\d+)\)\s*$')
_COURSE_LEVEL_RE = re.compile(r'^[A-Za-z]+(\d)')


def _mark_is_absent(mark):
    """Return True if *mark* indicates no real numeric result is available.

    Covers: None, blank, non-numeric strings (e.g. 'P', 'F'), and negative
    values (e.g. the -1 sentinel for a missing mark).
    """
    v = _numeric_mark(mark)
    return v is None or v < 0


def _mark_missing(mark):
    """Return True if a unit's mark cell is genuinely empty (blank or whitespace).

    Unlike _mark_is_absent this does NOT treat coded values such as 'P'/'F' as
    missing — those are entered data, not absent data.
    """
    return mark is None or (isinstance(mark, str) and not mark.strip())


def _norm_eid(v):
    """Normalise an emplid to a plain integer string for comparison."""
    try:
        return str(int(float(str(v).strip())))
    except (TypeError, ValueError):
        return str(v).strip()


def _append_code(existing, new_code):
    """Append *new_code* to *existing* output code, joining with '_' if needed."""
    return f"{existing}_{new_code}" if existing else new_code


_CODE_SPLIT_RE = re.compile(r'[\s,;/&|]+')

def _split_codes(value):
    """Return a frozenset of individual codes from a field that may hold several.

    Splits on any run of whitespace, comma, semicolon, slash, ampersand or pipe,
    so codes are picked up however they are separated, e.g. 'A, EA', 'EA AA' or
    'AA/EA' all yield the individual codes.
    """
    if not value:
        return frozenset()
    return frozenset(c for c in _CODE_SPLIT_RE.split(str(value).strip()) if c)

# 'AB' mit_circs code: a borderline overall-mark promotion extension.  A value
# such as 'AB -1%' beside ANY unit means the student's overall-mark boundary for
# borderline promotion is extended by that many percent (the magnitude of the
# number) — exactly like XB/BX in the BZ column, and additive with it.  It does
# NOT affect the unit it sits beside.  The magnitude must lie in this range.
AB_SHIFT_MIN = 0.5
AB_SHIFT_MAX = 2.0
_AB_CODE_RE = re.compile(r'(?<![A-Z])AB(?![A-Z])')   # 'AB' not embedded in another word
_AB_NUM_RE  = re.compile(r'-?\d+(?:\.\d+)?')          # first numeric value in the field


def _ab_promo_shift(mit_circs):
    """Return (shift, warning) for an 'AB' overall-mark promotion extension.

    shift   : the percent to extend the overall-mark promotion boundary by
              (magnitude of the number beside 'AB'), or 0.0 if no (valid) AB
              code is present.
    warning : a message if 'AB' is present but the number is missing or outside
              AB_SHIFT_MIN–AB_SHIFT_MAX (in which case shift is 0.0), else None.
    """
    if not mit_circs:
        return 0.0, None
    text = str(mit_circs)
    if not _AB_CODE_RE.search(text.upper()):
        return 0.0, None
    # Bare 'AB' (no number) is a pre-existing code, not a boundary relaxation —
    # leave it alone without warning.  Only an 'AB' carrying a numeric value is
    # treated as the relaxation, and only that value is range-checked.
    m = _AB_NUM_RE.search(text)
    if m is None:
        return 0.0, None
    shift = abs(float(m.group()))
    if not (AB_SHIFT_MIN <= shift <= AB_SHIFT_MAX):
        return 0.0, (f"'AB' shift {shift}% outside expected {AB_SHIFT_MIN}-{AB_SHIFT_MAX}% "
                     f"range in mit_circs {text!r}; not applied")
    return shift, None

# Classyears where deferrals (EA) and resits apply.
_DEFERRAL_CLASSYEARS = frozenset({'1', '1m', '2', '2m'})

# Classyears where L3/L4 credit tallies are computed.
_LEVEL_CREDIT_CLASSYEARS = frozenset({'31', '31m', '32', '32m', '4', '4m'})

# EN codes indicating a mark carried forward from a previous attempt.
_CARRIED_EN_CODES = frozenset({'L1C', 'L2C', 'L3C'})

# Overall-mark year weights for physics (non-M+P) programmes.
# Keys are year numbers (1–4); values are base weights (renormalised when years are missing).
_BSC_WEIGHTS          = {1: 10,  2: 30,  3: 60}           # BSc and MPhys Y1–Y3
_MPHYS_WEIGHTS        = {1:  6,  2: 19,  3: 37.5, 4: 37.5}  # standard MPhys Y4
_MPHYS_ABROAD_WEIGHTS = {1:  8,  2: 23,  3: 23,   4: 46}    # MPhys with Y3 study abroad


#==================================================================================
# Functions!

def _parse_module(module):
    """Split 'PHYS10180 (20)' into ('PHYS10180', 20).

    Returns (coursename, credits) where credits is an int, or
    (module, None) if the string doesn't match the expected pattern.
    """
    if module is None:
        return None, None
    m = _MODULE_RE.match(str(module))
    if m:
        return m.group(1).strip(), int(m.group(2))
    return module, None


def _course_level(coursename):
    """Return the level digit (int) from a module code like 'PHYS30471', or None.

    The level is the first digit after the leading letters, e.g. PHYS3xxxx → 3.
    """
    if not coursename:
        return None
    m = _COURSE_LEVEL_RE.match(coursename)
    return int(m.group(1)) if m else None


_MARK_NUM_RE = re.compile(r'^\s*([\d.]+)\s*([A-Za-z]*)\s*$')

def _numeric_mark(value):
    """Return the numeric part of a mark value, stripping any trailing letter suffix.

    Handles plain floats and values like '35C' (compensated) or '30R' (resit).
    Returns None if no numeric value can be extracted.
    """
    if value is None:
        return None
    m = _MARK_NUM_RE.match(str(value).strip())
    if m:
        try:
            return float(m.group(1))
        except ValueError:
            return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None

def _is_mc_excluded(output_code):
    """True if *output_code* marks a mitigating-circumstances exclusion.

    A unit excluded by mitigating circumstances (an AA in mit_circs) carries
    either a standalone 'X' action code (e.g. 'XN_X', 'AA_EA_X') or, for the pure
    missed-exam case (XN + AA only), the single code 'XL'.  A deferral ('R1',
    'XL_R1', ...) or a carried mark ('L1C', 'L2C', 'L3C') is also excluded from
    the year mark but does NOT represent an MC exclusion, so those marks are not
    highlighted green.  Splitting on '_' keeps 'XL_R1' (a deferral) and 'XN'
    (missed exam) from matching, while 'XN_X', 'AA_EA_X' and a lone 'XL' do.
    """
    tokens = (output_code or '').split('_')
    return 'X' in tokens or tokens == ['XL']


def _raw_input_codes(unit):
    """Display string of a unit's raw input codes (EN + mit_circs).

    The 'R1' / 'R2' resit markers are converted to '1st att.' / '2nd att.' and
    placed after any other codes; every other code is kept, in order.  Used for
    completed / withdrawn / interrupted / manual / intercalating students, who are
    not reassessed and so carry no computed action codes.
    str.split() handles the stray non-breaking spaces seen in mit_circs values.
    """
    other = []
    att   = []   # 'R1'/'R2' → '1st att.'/'2nd att.', shown after any other codes
    for field in (unit.en, unit.mit_circs):
        if not field:
            continue
        for tok in str(field).split():
            u = tok.upper()
            if u == 'R1':
                att.append('1st att.')
            elif u == 'R2':
                att.append('2nd att.')
            else:
                other.append(tok)
    return ' '.join(other + att) or None


def _mark_suffix(value):
    """Return the letter suffix of a mark value in upper case, or '' if none.

    E.g. '30R' → 'R', '35C' → 'C', '40' → ''.
    """
    if value is None:
        return ''
    m = _MARK_NUM_RE.match(str(value).strip())
    return m.group(2).upper() if m else ''


def _mark_accepted(value):
    """True if a unit mark counts as a pass.

    A mark passes when its numeric value exceeds PASS_MARK, or when it carries a
    'C' (compensated) or 'R' (passed resit) accept-suffix — the board has accepted
    such a mark, so the unit counts as passed even below PASS_MARK.
    """
    num = _numeric_mark(value)
    return (num is not None and num > PASS_MARK) or _mark_suffix(value) in ('C', 'R')


def _order_units(units):
    """Return *units* ordered for the output grid: the lab unit(s) first, then the
    project / dissertation unit(s), then every other unit in its original input order.

    Python's sort is stable, so units within each group keep their input-file order —
    in particular 'the rest' are left exactly as they came in.  Matching is on the
    bare course code (coursename), so empty/unparsed slots fall through to the rest.
    """
    def _group(u):
        code = (u.coursename or u.module or '').strip()
        if code in MUST_PASS_LAB:
            return 0
        if code in PROJECT_MODULES:
            return 1
        return 2
    return sorted(units, key=_group)


class UnitInfo:
    """Data for a single unit (module) for one student."""
    __slots__ = ('unit_name', 'module', 'coursename', 'credits', 'mark', 'en', 'mit_circs',
                 'passed', 'excluded', 'output_code', 'capped')

    def __init__(self, unit_name, module, mark, en, mit_circs):
        self.unit_name   = unit_name              # 'Unit 1', 'Unit 2', etc.
        self.module      = module                 # raw cell value, e.g. 'PHYS10180 (20)'
        self.coursename, self.credits = _parse_module(module)
        self.mark        = mark                   # unit mark (float) or None
        self.en          = en                     # EN flag (str) or None
        self.mit_circs   = mit_circs              # mitigating circumstances code (str) or None
        self.passed      = None                   # True/False once exclude_units() runs; None = no mark
        self.excluded    = False                  # True if excluded from year mark
        self.output_code = None                   # code(s) written to the output code column
        self.capped      = False                  # True if mark is capped at 30.0 for year mark (R2 attempt)

    def __repr__(self):
        return (f'UnitInfo({self.unit_name}, coursename={self.coursename!r}, '
                f'credits={self.credits}, mark={self.mark}, passed={self.passed}, '
                f'excluded={self.excluded}, output_code={self.output_code!r})')


class StudentInfo:
    """All data for a single student, read from one row of the exam grid."""
    __slots__ = (
        'emplid', 'name', 'id_no', 'uf', 'mc', 'bz',
        'admit_term', 'entry_type', 'psi', 'plan', 'is_mphys_track', 'is_study_abroad',
        'is_pp', 'is_abroad',
        'AS_code', 'RFYR', 'RYOA', 'RYIA', 'COMP',
        'units_passed', 'award', 'classification',
        'units', 'trailing',
        'yearmark',
        'credits_taken', 'credits_passed', 'credits_excluded', 'credits_deferred', 'creds_passed_taken',
        'excluded_idx', 'excluded_courses',
        'failed_idx', 'failed_courses',
        'deferred_idx', 'deferred_courses',
        'some_unit_under_30',
        'zone_idx', 'zone_courses',
        'compensated_idx', 'compensated_courses', 'credits_compensated',
        'referred_idx', 'referred_courses',
        'resits',
        'fail', 'fail_reason',
        'status',
        'phys_yearmark', 'math_yearmark',
        'phys1', 'phys2', 'phys3',
        'credits_l3', 'credits_l4', 'l3_l4_creds_passed',
        'credits_l3_first', 'credits_l3_upper2', 'credits_l3_lower2', 'credits_l3_third',
        'credits_at_first', 'credits_at_upper2', 'credits_at_lower2', 'credits_at_third',
        'overall',
        'project_mark', 'project_creds',
        'deg_class_alg', 'deg_class_actual',
        'borderline_for', 'deg_class_rev', 'deg_class_rev_detail',
        'y3creds_below40_not_excl', 'y3creds_below40_excl', 'y3creds_l4_passed',
        'y3creds_l4_taken',
        'y3creds_below40',
        'l3_l4_credits_failed', 'credits_passed_y3y4',
        'y3_creds_failed_str', 'l4_creds_y3y4_str',
        'cf_flags',
    )

    def __init__(self):
        self.emplid         = None
        self.name           = None
        self.id_no          = None   # sequential row number in the grid
        self.uf             = None
        self.mc             = None
        self.bz             = None
        self.admit_term     = None
        self.entry_type     = None
        self.psi            = None
        self.plan           = None   # degree programme, e.g. 'MPhys(Hons) Physics'
        self.is_mphys_track  = False  # True if plan is MPhys or MMath (4-year course)
        self.is_study_abroad = False  # True if year-abroad (plan has 'study', or in abroad_list)
        self.is_pp           = False  # True if student has a professional placement unit
        self.is_abroad       = False  # True if student has a year abroad unit
        self.AS_code         = None   # raw 'AS Code' column value (e.g. 'ACTV', 'RFYR', 'EXIT')
        self.RFYR            = False  # True if AS_code is RFYR (repeat first year → Interrupt)
        self.RYOA            = False  # True if AS_code is RYOA (repeat year out abroad → Interrupt)
        self.RYIA            = False  # True if AS_code is RYIA (repeat year in attendance → Interrupt)
        self.COMP            = False  # True if AS_code is COMP (completed); recorded only, no special handling
        self.units_passed   = None
        self.award          = None
        self.classification = None
        self.units          = []     # list of UnitInfo, one per unit column block
        self.trailing       = {}     # trailing columns: normalised name -> value
        self.yearmark           = None   # credit-weighted average of unit marks
        self.credits_taken      = None   # total credits with a mark
        self.credits_passed     = None   # credits where mark > PASS_MARK
        self.credits_excluded   = 0      # credits excluded from calculation (populated later)
        self.credits_deferred   = 0      # credits from deferred (EA) units (assumed passed)
        self.creds_passed_taken = None   # formatted string for output, e.g. '120 / 120'
        self.excluded_idx       = []     # indices into self.units of excluded units
        self.excluded_courses   = []     # coursenames of excluded units
        self.failed_idx         = []     # indices into self.units of failed units
        self.failed_courses     = []     # coursenames of failed units
        self.deferred_idx       = []     # indices into self.units of deferred (EA) units
        self.deferred_courses   = []     # coursenames of deferred units
        self.some_unit_under_30  = False  # True if any failed unit is below MIN_MARK (30%)
        self.zone_idx            = []     # indices of units in compensation zone (MIN_MARK < mark <= PASS_MARK)
        self.zone_courses        = []     # coursenames of zone units
        self.compensated_idx     = []     # indices of units compensated (failed but allowed to count)
        self.compensated_courses = []     # coursenames of compensated units
        self.credits_compensated = 0      # total credits compensated
        self.referred_idx        = []     # indices of units referred (R2 resit)
        self.referred_courses    = []     # coursenames of referred units
        self.resits              = None   # deferred/resit courses for output, e.g. 'PHYS10071[1] / PHYS10101[1]'
        self.fail                = False  # True if student cannot progress
        self.fail_reason         = ''     # short description of why student failed
        self.status              = None   # set once by calc_status(): 'ACTV', 'A/D', 'REVW', 'FAIL'
        self.phys_yearmark       = None   # M+P only: credit-weighted average of non-MATH units
        self.math_yearmark       = None   # M+P only: credit-weighted average of MATH units
        self.phys1               = None   # previous Y1 mark (from L1CM trailing column)
        self.phys2               = None   # previous Y2 mark (from L2CM trailing column)
        self.phys3               = None   # previous Y3 mark (from L3CM trailing column)
        self.credits_l3          = 0      # credits passed at level 3
        self.credits_l4          = 0      # credits passed at level 4 (incl. level 6, excl. level 5)
        self.l3_l4_creds_passed  = None   # formatted string, e.g. '80 + 40 = 120'
        self.credits_l3_first    = 0      # L3 credits (incl. excluded) with mark >= BOUNDARY_FIRST
        self.credits_l3_upper2   = 0      # L3 credits (incl. excluded) with mark >= BOUNDARY_UPPER2
        self.credits_l3_lower2   = 0      # L3 credits (incl. excluded) with mark >= BOUNDARY_LOWER2
        self.credits_l3_third    = 0      # L3 credits (incl. excluded) with mark >= BOUNDARY_THIRD
        self.credits_at_first  = 0    # all current-year credits (incl. excluded) with mark >= BOUNDARY_FIRST
        self.credits_at_upper2 = 0    # all current-year credits (incl. excluded) with mark >= BOUNDARY_UPPER2
        self.credits_at_lower2 = 0    # all current-year credits (incl. excluded) with mark >= BOUNDARY_LOWER2
        self.credits_at_third  = 0    # all current-year credits (incl. excluded) with mark >= BOUNDARY_THIRD
        self.overall             = None   # weighted overall mark across years
        self.project_mark        = None   # credit-weighted average project mark (rounded int), or None
        self.project_creds       = 0      # total project credits
        self.deg_class_alg       = None   # algorithmic degree classification, e.g. 'BSc 2.1'
        self.deg_class_actual    = None   # actual degree classification (board may override)
        self.borderline_for      = None   # class student is borderline for, e.g. '1', '2.1'
        self.deg_class_rev       = None   # promotion note: 'P(A)' or 'P(B)', else None
        self.deg_class_rev_detail = None  # not shown: BSc 'CR' borderline breakdown, e.g. 'CR/marks'
        self.y3creds_below40_not_excl   = None   # Y4 only: Y3 credits failed, not excluded (from Y3 credits sheet)
        self.y3creds_below40_excl       = None   # Y4 only: Y3 credits failed with MCs, excluded
        self.y3creds_l4_passed          = None   # Y4 only: L4 credits passed in Y3
        self.y3creds_l4_taken           = None   # Y4 only: L4+ credits taken in Y3 (from Y3 credits sheet)
        self.y3creds_below40            = None   # Y4 only: total Y3 credits below 40 (not_excl + excl)
        self.l3_l4_credits_failed       = None   # Y4 only: y3creds_below40 + L3+ credits failed in current Y4 grid
        self.credits_passed_y3y4        = None   # Y4 only: any-level credits passed over Y3+Y4 = 240 - l3_l4_credits_failed; used for the award threshold
        self.y3_creds_failed_str        = None   # Y4 only: "not_excl/excl" for 'Y3 creds failed w/wo MCs' column
        self.l4_creds_y3y4_str          = None   # Y4 only: "y3+y4=total" for 'L4 creds passed Y3+Y4' column
        self.cf_flags                   = ''     # carry-forward notes from CF_FLAG_FILE; blank if not matched

    def exclude_units(self, classyear=None):
        """Set credits_taken/passed/excluded, creds_passed_taken, and unit flags.

        Exclusion rules applied here:
          EA in mit_circs → deferral; only actioned for years 1/2
                            (classyear in _DEFERRAL_CLASSYEARS). Excluded from year
                            mark, treated as passed for outcome checks, but NOT
                            counted in credits_passed.
          CA in mit_circs → deferral (years 1/2 only); excluded from year mark,
                            treated as passed for outcome checks, but NOT counted
                            in credits_passed.
          AA in mit_circs → excluded from year mark, treated as passed for progression,
                            output_code set to 'X'.
          AA + a deferral code (EA or CA) both present → early grids carry both an
                            exclusion and a deferral code; the exam board later resolves
                            them to one. Until then they are resolved here by result: a
                            PASSED unit is treated as 'AA', a FAILED unit as the deferral
                            code (the deferral itself only applies in years 1/2; in years
                            3/4 a combined code falls back to 'AA'). Both input codes are
                            copied to the output column (before the action code) for info.
          credits_passed is determined by mark (> PASS_MARK) or an accepted-mark
          suffix ('C' compensated / 'R' passed resit), regardless of exclusion
          codes.  Excluded units whose mark does not pass instead count toward the
          progression check via credits_deferred.
        """
        taken            = 0
        passed           = 0
        excluded         = 0
        deferred_creds   = 0
        excluded_idx     = []
        excluded_courses = []
        failed_idx       = []
        failed_courses   = []
        deferred_idx     = []
        deferred_courses = []

        for idx, unit in enumerate(self.units):
            if unit.module is None or unit.credits is None:
                continue                             # empty slot

            taken     += unit.credits
            mit_codes  = _split_codes(unit.mit_circs)
            en_codes   = _split_codes(unit.en)
            used_mit   = set()                       # mit codes consumed by a rule
            used_en    = set()                       # EN codes consumed by a rule
            action_codes = []                        # codes generated by rules (X, R1, ...), output after input codes
            coursename = unit.coursename or unit.module

            # --- Resolve AA / EA / CA mitigation (incl. combined AA+EA, AA+CA cases) ---
            # Early grids may carry an exclusion code (AA) AND a deferral code (EA or CA)
            # together; the board later reduces them to one. Until then resolve by result:
            # a PASSED unit is treated as 'AA' (excluded), a FAILED unit as the deferral
            # code. Deferrals only apply in years 1/2; in years 3/4 a combined code falls
            # back to 'AA'. Both input codes are copied to the output (before the action
            # code) for info.
            has_aa        = 'AA' in mit_codes
            has_ea        = 'EA' in mit_codes
            has_ca        = 'CA' in mit_codes
            defer_code    = 'EA' if has_ea else ('CA' if has_ca else None)  # deferral code present (EA preferred)
            both_aa_defer = has_aa and defer_code is not None
            # course_passed honours the 'C'/'R' accept-suffix (see _mark_accepted).
            course_passed = _mark_accepted(unit.mark)

            if both_aa_defer:
                eff_code = 'AA' if course_passed else defer_code
            elif defer_code:
                eff_code = defer_code
            elif has_aa:
                eff_code = 'AA'
            else:
                eff_code = None
            if eff_code in ('EA', 'CA') and classyear not in _DEFERRAL_CLASSYEARS:
                # No deferral outside years 1/2: a combined code becomes AA, a lone deferral is left.
                eff_code = 'AA' if has_aa else None

            # --- EA / CA deferral (years 1/2 only) ---
            if eff_code in ('EA', 'CA'):
                unit.excluded = True
                unit.passed   = True           # treated as passed for outcome checks
                if not both_aa_defer:
                    used_mit.add(eff_code)     # lone deferral: represented by the action code, not echoed
                # (combined: leave AA and the deferral code unconsumed so both echo as input codes)
                if eff_code == 'EA' and 'XN' in en_codes:   # missed exam → XL_R1
                    action_codes.append('XL_R1')
                    used_en.add('XN')
                else:
                    action_codes.append('R1')
                excluded += unit.credits
                excluded_idx.append(idx);  excluded_courses.append(coursename)
                deferred_idx.append(idx);  deferred_courses.append(coursename)

            # --- AA exclusion (also the combined AA+EA / AA+CA case when resolved to AA) ---
            elif eff_code == 'AA':
                unit.excluded    = True
                unit.passed      = True           # treated as passed for outcome checks
                if not both_aa_defer:
                    used_mit.add('AA')         # lone AA: represented by 'X', not echoed
                # (combined: leave AA and the deferral code unconsumed so both echo as input codes)
                # Pure missed-exam exclusion — XN + AA and nothing else (no AB or
                # other codes; only arises in years 3/4) — is shown as a single 'XL'
                # code instead of 'XN_X'.  Display only: the XN echo is consumed and
                # the 'X' action swapped for 'XL'; no exclusion/credit/outcome change.
                if en_codes == frozenset({'XN'}) and mit_codes == frozenset({'AA'}):
                    used_en.add('XN')
                    action_codes.append('XL')
                else:
                    action_codes.append('X')
                excluded += unit.credits
                excluded_idx.append(idx);  excluded_courses.append(coursename)

            # --- L1C/L2C/L3C (carried mark): exclude from year average, treat as passed ---
            elif en_codes & _CARRIED_EN_CODES:
                unit.excluded = True
                unit.passed   = True
                excluded += unit.credits
                excluded_idx.append(idx);  excluded_courses.append(coursename)
                # carried code not added to used_en, so it copies through to the output code column

            # --- XN (missed exam): counts as failed, mark still used in year mark average ---
            elif 'XN' in en_codes:
                unit.passed = False
                failed_idx.append(idx);  failed_courses.append(coursename)
                # XN not added to used_en, so it copies through to the output code column

            # --- normal pass/fail (including no mark, which counts as failed) ---
            else:
                unit.passed = course_passed
                if not unit.passed:
                    failed_idx.append(idx);  failed_courses.append(coursename)

            # --- credit accumulation: purely mark-based for credits_passed ---
            # Excluded units whose mark doesn't pass still count for the progression check.
            if course_passed:
                passed += unit.credits
            elif unit.excluded:
                deferred_creds += unit.credits

            # --- assemble output_code: all input codes first, then any action codes ---
            # Input codes are the unprocessed EN and mit_circs codes carried through for
            # info; action codes (X, R1, ...) generated by the rules above come after,
            # e.g. 'XN_X' not 'X_XN', 'AA_EA_R1' not 'R1_AA_EA'.
            input_codes = []
            att_codes   = []   # input-EN 'R1'/'R2' → '1st att.'/'2nd att.', shown after any other codes
            for code in sorted(en_codes - used_en):
                if code == 'R2':
                    # A 'C' (compensated) or 'R' (passed resit) suffix means the
                    # 2nd-attempt mark has been accepted: it stands uncapped and the
                    # '2nd att.' note is redundant, so omit it.  Otherwise label the
                    # attempt and cap the mark at 30 in the year mark.
                    if _mark_suffix(unit.mark) not in ('C', 'R'):
                        att_codes.append('2nd att.')
                        unit.capped = True
                elif code == 'R1':
                    att_codes.append('1st att.')   # 1st-attempt resit: treated normally, labelled for clarity
                else:
                    input_codes.append(code)
            input_codes.extend(sorted(mit_codes - used_mit))   # mit codes not consumed by a rule
            parts = input_codes + att_codes + action_codes
            unit.output_code = '_'.join(parts) if parts else None

        self.credits_taken      = taken
        self.credits_passed     = passed
        self.credits_excluded   = excluded
        self.credits_deferred   = deferred_creds
        self.creds_passed_taken = f"{passed} / {taken}"
        self.excluded_idx       = excluded_idx
        self.excluded_courses   = excluded_courses
        self.failed_idx         = failed_idx
        self.failed_courses     = failed_courses
        self.deferred_idx       = deferred_idx
        self.deferred_courses   = deferred_courses
        self.resits             = ' / '.join(f"{c}[1]" for c in deferred_courses if c) or None

    def calc_yearmark(self, classyear=None):
        """Set self.yearmark to the credit-weighted mean of all unit marks.

        For M+P classyears also sets self.phys_yearmark (non-MATH units) and
        self.math_yearmark (units whose coursename contains 'MATH').
        Units with a missing mark or missing credits are skipped.
        Results are rounded to 1 decimal place; '-1' if no valid units found.
        """
        def _round1dp(weighted, credits):
            return math.floor(weighted / credits * 10 + 0.5) / 10 if credits > 0 else '-1'

        total_credits  = 0
        weighted_marks = 0.0
        phys_credits   = 0
        phys_weighted  = 0.0
        math_credits   = 0
        math_weighted  = 0.0
        mp = classyear is not None and classyear.endswith('m')

        for unit in self.units:
            if unit.excluded:
                continue
            if unit.mark is None or unit.credits is None:
                continue
            mark = _numeric_mark(unit.mark)
            if mark is None:
                continue
            calc_mark = min(mark, 30.0) if unit.capped else mark
            weighted_marks += calc_mark * unit.credits
            total_credits  += unit.credits
            if mp:
                if (unit.coursename or '').startswith('MATH'):
                    math_weighted += calc_mark * unit.credits
                    math_credits  += unit.credits
                else:
                    phys_weighted += calc_mark * unit.credits
                    phys_credits  += unit.credits

        self.yearmark = _round1dp(weighted_marks, total_credits)
        if mp:
            self.phys_yearmark = _round1dp(phys_weighted, phys_credits)
            self.math_yearmark = _round1dp(math_weighted, math_credits)

    def calc_status(self, classyear):
        """Set self.status once, reading flags set by earlier methods.

        Priority order:
          FAIL  — self.fail is True
          REVW  — one or more referrals (self.referred_idx non-empty), with or without deferrals
          A/D   — one or more deferrals, no referrals (self.deferred_idx non-empty)
          ACTV  — default: active, fine to progress

        Y2 MPhys/MMath additional yearmark check (non-FAIL students on 4-year course).
        When yearmark <= MPHYS_PROGRESS_MARK (borderline or below), by referral/deferral mix:
          deferrals only (no referrals)        → A/D (kept on MPhys/MMath, even if below review mark)
          any referrals (with or without deferrals):
            borderline (> MPHYS_REVIEW_MARK)   → REVW R/X
            below      (<= MPHYS_REVIEW_MARK)  → REVW (BSc)
          neither (ACTV):
            borderline (> MPHYS_REVIEW_MARK)   → R/X
            below      (<= MPHYS_REVIEW_MARK)  → BSc
        A yearmark > MPHYS_PROGRESS_MARK leaves the status unchanged.
        """
        if classyear in FINAL_CLASSYEARS:
            return
        if self.fail:
            self.status = 'FAIL'
            self.resits = None
        elif self.referred_idx:
            self.status = 'REVW'
        elif self.deferred_idx:
            self.status = 'A/D'
        else:
            self.status = 'ACTV'

        # Y2 MPhys/MMath yearmark progression check (4-year course, non-FAIL students).
        if (classyear in ('2', '2m')
                and self.is_mphys_track
                and self.status != 'FAIL'):
            ym = self.yearmark
            if ym is None or ym > MPHYS_PROGRESS_MARK:
                pass  # no yearmark, or clears the progression mark — leave status as-is
            else:
                # borderline = MPHYS_REVIEW_MARK < ym <= MPHYS_PROGRESS_MARK; else below.
                borderline = ym > MPHYS_REVIEW_MARK
                if self.referred_idx:
                    # Any referrals (with or without deferrals)
                    self.status = 'REVW R/X' if borderline else 'REVW (BSc)'
                elif self.deferred_idx:
                    self.status = 'A/D'    # deferrals only → stays on MPhys/MMath
                else:
                    # No deferrals or referrals (ACTV)
                    self.status = 'R/X' if borderline else 'BSc'

        # Y2 CertHE: a failing Y2 student is awarded a Certificate of Higher
        # Education.  Status becomes 'FAIL (CertHE)' rather than plain 'FAIL'.
        if self.status == 'FAIL' and self.y2_certhe_eligible(classyear):
            self.status = 'FAIL (CertHE)'

    def y2_certhe_eligible(self, classyear):
        """True if a Y2 student qualifies for a Certificate of Higher Education.

        Requires that the student is NOT a direct entrant (L1CM is a real mark,
        not the -1 'no Year-1 mark' sentinel) and is NOT carrying credits
        (registered for exactly 120 credits in Y2).
        """
        if classyear not in ('2', '2m') or self.credits_taken != 120:
            return False
        l1cm = _numeric_mark(self.phys1)
        return l1cm is not None and l1cm >= 0

    def calc_bsc_class_y3mphys(self, classyear):
        """Check Y3→Y4 progression for MPhys/MMath students (classyear 31/31m).

        Progression requires ALL of:
          yearmark       > MPHYS_Y3_PROG_YEARMARK
          overall        > MPHYS_Y3_PROG_OVERALL
          credits_passed >= MPHYS_Y3_PROG_CREDITS
          no unresittable lab (MUST_PASS_LAB) failure
          (MMath only) phys and maths yearmarks > MPHYS_Y3_PROG_PHYS_MATH_MARK

        Students who already have status 'FAIL' are left unchanged.  Students
        who fail any criterion are treated as BSc candidates: the equivalent
        BSc classification is computed and written to self.status and the
        trailing 'Award' column.  Status is suffixed with ' (CR)' (and the
        Award left blank for board review) when the student is in a borderline
        zone; otherwise Status and Award both hold the clean-cut BSc class.

        self.fail_reason records which criteria sent the student to BSc
        consideration (credits, average, overall, Phys/Maths averages, or a
        failed compulsory lab), shown in the 'Fail reason' output column.
        """
        if self.status == 'FAIL':
            return

        try:
            ym = float(self.yearmark)
        except (TypeError, ValueError):
            ym = -1.0
        try:
            ov = float(self.overall)
        except (TypeError, ValueError):
            ov = -1.0

        phys_ok = math_ok = True
        if classyear.endswith('m'):
            try:
                phys_ok = float(self.phys_yearmark) > MPHYS_Y3_PROG_PHYS_MATH_MARK
            except (TypeError, ValueError):
                phys_ok = False
            try:
                math_ok = float(self.math_yearmark) > MPHYS_Y3_PROG_PHYS_MATH_MARK
            except (TypeError, ValueError):
                math_ok = False

        # A lab (MUST_PASS_LAB, any year/semester) that is failed and not being resat
        # cannot be carried into Y4 — the student cannot continue MPhys/MMath without
        # it.  Using the lab set (not MUST_PASS) keeps placeholder project units out.
        lab_failed = any(
            (self.units[idx].coursename or self.units[idx].module or '') in MUST_PASS_LAB
            for idx in self.failed_idx
            if idx not in self.referred_idx
        )

        if (ym > MPHYS_Y3_PROG_YEARMARK
                and ov > MPHYS_Y3_PROG_OVERALL
                and (self.credits_passed or 0) >= MPHYS_Y3_PROG_CREDITS
                and phys_ok and math_ok
                and not lab_failed):
            return   # meets all criteria — progress to Y4 unchanged

        # Failed at least one criterion: compute BSc classification.
        prior_reason = self.fail_reason   # e.g. 'Resit failed lab' from calc_referrals
        bsc_cy = '32m' if classyear.endswith('m') else '32'
        self.calc_project_mark(bsc_cy)
        self.calc_deg_class(bsc_cy)

        # Record why the student dropped to BSc consideration.  This overrides the
        # BSc-Fail reasons set by calc_deg_class so the column reads as the
        # MPhys/MMath progression miss (like the Fail reason shown in other years).
        reasons = []
        if prior_reason:
            reasons.append(prior_reason)
        creds = self.credits_passed or 0
        if creds < MPHYS_Y3_PROG_CREDITS:
            reasons.append(f'< {MPHYS_Y3_PROG_CREDITS} credits ({creds})')
        if not (ym > MPHYS_Y3_PROG_YEARMARK):
            reasons.append(f'yearmark < {int(MPHYS_Y3_PROG_YEARMARK + 0.5)}')
        if not (ov > MPHYS_Y3_PROG_OVERALL):
            reasons.append(f'overallmark < {int(MPHYS_Y3_PROG_OVERALL + 0.5)}')
        if not phys_ok:
            reasons.append(f'Phys average < {int(MPHYS_Y3_PROG_PHYS_MATH_MARK + 0.5)}')
        if not math_ok:
            reasons.append(f'Maths average < {int(MPHYS_Y3_PROG_PHYS_MATH_MARK + 0.5)}')
        if lab_failed and not any('lab' in r.lower() for r in reasons):
            reasons.append('Failed lab')
        self.fail_reason = ' / '.join(reasons)

        bsc_class = self.deg_class_actual
        if bsc_class is None:
            return

        if self.borderline_for is not None:
            self.status = f'{bsc_class} (CR)'
            self.trailing['Award'] = None   # borderline: leave Award blank for board review
        else:
            self.status = bsc_class
            self.trailing['Award'] = bsc_class

    def apply_special_status(self, classyear):
        """Override classification/status for completed, interrupted, manual or
        withdrawn students.

        A student flagged self.COMP (AS code 'COMP') has already completed and is
        not reassessed: 'Completed' is written to status, deg_class_alg and
        deg_class_actual.  Otherwise, if the student's emplid is in interrupt_list,
        manual_list, or withdrawn_list:
          - Progressing years: sets self.status to the label.
          - Final years: sets deg_class_alg and deg_class_actual to the label.
        In all cases clears resits, fail_reason, deg_class_rev (and its
        deg_class_rev_detail breakdown), borderline_for,
        and replaces every unit's output_code with the raw EN + mit_circs values
        from the input (no computed action codes).  The credits and averages
        computed earlier (creds_passed_taken, yearmark, overall, ...) are kept.
        Returns True if the student matched, False otherwise.
        """
        eid = _norm_eid(self.emplid)
        _norm = lambda lst: {_norm_eid(e) for e in lst}
        if   self.COMP:
            label = 'Completed'
        elif eid in _norm(interrupt_list):
            label = 'Interrupt'
        elif eid in _norm(manual_list):
            label = 'Manual'
        elif eid in _norm(withdrawn_list):
            label = 'Withdrawn'
        else:
            return False

        if label == 'Completed':
            # Already completed: mark every outcome column 'Completed'.  The marks,
            # raw output codes and the credits/averages set earlier are retained.
            self.status           = label
            self.deg_class_alg    = label
            self.deg_class_actual = label
            self.deg_class_rev    = None
            self.deg_class_rev_detail = None
            self.borderline_for   = None
        elif classyear in FINAL_CLASSYEARS:
            self.deg_class_alg    = label
            self.deg_class_actual = label
            self.deg_class_rev    = None
            self.deg_class_rev_detail = None
            self.borderline_for   = None
        else:
            self.status = label
            # A withdrawn Y2 student may still qualify for a CertHE exit award.
            if label == 'Withdrawn' and self.y2_certhe_eligible(classyear):
                self.status = 'Withdrawn (CertHE)'

        self.resits   = None
        if label == 'Withdrawn':
            notes_text = self.cf_flags or self.trailing.get('Notes') or ''
            self.fail_reason = ('' if 'withdrawn' in str(notes_text).lower()
                                else 'Withdrawn')
        else:
            self.fail_reason = ''

        for unit in self.units:
            unit.output_code = _raw_input_codes(unit)

        # Non-awarded student with no current-year marks: the year mark is the
        # '-1' sentinel, so blank the overall too (don't show a prior-years
        # synthesis for someone who isn't being classified this year).
        if _mark_is_absent(self.yearmark):
            self.overall = '-1'

        return True

    def detect_intercal(self, classyear):
        """Detect professional placement (PP) or year-abroad students with absent marks.

        Scans every unit for PP_MODULES (PHYS30810/PHYS40810) and ABROAD_MODULES
        (PHYS31000/PHYS41000), setting self.is_pp and self.is_abroad regardless of
        whether marks are present.

        If any matched unit has an absent mark (blank, -1, or non-numeric such as
        'P'/'F'), applies 'Intercal' status and performs the same output blanking
        as apply_special_status (resits/fail_reason cleared, unit output codes
        replaced by raw EN + mit_circs values).
        """
        pp_absent     = False
        abroad_absent = False

        for unit in self.units:
            code = (unit.coursename or '').strip()
            if code in PP_MODULES:
                self.is_pp = True
                if _mark_is_absent(unit.mark):
                    pp_absent = True
            if code in ABROAD_MODULES:
                self.is_abroad = True
                if _mark_is_absent(unit.mark):
                    abroad_absent = True

        if not (pp_absent or abroad_absent):
            return   # marks are present — process normally

        if classyear in FINAL_CLASSYEARS:
            self.deg_class_alg    = 'Intercal'
            self.deg_class_actual = 'Intercal'
            self.deg_class_rev    = None
            self.deg_class_rev_detail = None
            self.borderline_for   = None
        else:
            self.status = 'Intercal'

        self.resits      = None
        self.fail_reason = ''

        for unit in self.units:
            unit.output_code = _raw_input_codes(unit)

        # No current-year marks (year mark = '-1' sentinel): blank the overall too
        # rather than display a prior-years synthesis for an intercalating student.
        if _mark_is_absent(self.yearmark):
            self.overall = '-1'

    def calc_level_credits(self):
        """Count passed credits at level 3 and level 4 (level 6 treated as level 4; level 5 excluded).

        Also tallies, at each degree-class boundary (mark >= boundary), the L3-only
        credits (credits_l3_*) and ALL current-year credits of any level (credits_at_*,
        used for borderline promotion).  All counts include excluded units — only the
        numeric mark is tested.
        Sets credits_l3, credits_l4, l3_l4_creds_passed, credits_l3_* and credits_at_*.
        """
        l3 = 0
        l4 = 0
        l3_first  = 0
        l3_upper2 = 0
        l3_lower2 = 0
        l3_third  = 0
        at_first  = 0
        at_upper2 = 0
        at_lower2 = 0
        at_third  = 0
        for unit in self.units:
            if unit.credits is None or unit.module is None:
                continue
            mark_num = _numeric_mark(unit.mark)
            if mark_num is None:
                continue
            eff = mark_num
            level = _course_level(unit.coursename or unit.module)
            if level == 3:
                if eff > PASS_MARK:
                    l3 += unit.credits
                if eff >= BOUNDARY_FIRST:
                    l3_first += unit.credits
                if eff >= BOUNDARY_UPPER2:
                    l3_upper2 += unit.credits
                if eff >= BOUNDARY_LOWER2:
                    l3_lower2 += unit.credits
                if eff >= BOUNDARY_THIRD:
                    l3_third += unit.credits
            elif level in (4, 6):
                if eff > PASS_MARK:
                    l4 += unit.credits
            # Credits at each class boundary count ALL current-year units (any
            # level, incl. excluded) — used for borderline promotion (Alg A/B).
            if eff >= BOUNDARY_FIRST:
                at_first += unit.credits
            if eff >= BOUNDARY_UPPER2:
                at_upper2 += unit.credits
            if eff >= BOUNDARY_LOWER2:
                at_lower2 += unit.credits
            if eff >= BOUNDARY_THIRD:
                at_third += unit.credits
        self.credits_l3           = l3
        self.credits_l4           = l4
        self.l3_l4_creds_passed   = f"{l3} + {l4} = {l3 + l4}"
        self.credits_l3_first     = l3_first
        self.credits_l3_upper2    = l3_upper2
        self.credits_l3_lower2    = l3_lower2
        self.credits_l3_third     = l3_third
        self.credits_at_first  = at_first
        self.credits_at_upper2 = at_upper2
        self.credits_at_lower2 = at_lower2
        self.credits_at_third  = at_third

    def calc_overall(self, classyear):
        """Set self.overall to the weighted overall mark across available year marks.

        Physics (non-M+P) base weights:
          BSc / MPhys Y1–Y3 (_BSC_WEIGHTS):       Y1=10, Y2=30, Y3=60
          Standard MPhys Y4 (_MPHYS_WEIGHTS):      Y1=6,  Y2=19, Y3=37.5, Y4=37.5
          MPhys Y3 study abroad (_MPHYS_ABROAD_WEIGHTS): Y1=8, Y2=23, Y3=23, Y4=46

        Missing years (None or the '-1' sentinel) are dropped and the remaining
        weights renormalised so their ratios are preserved.  Sets self.overall to
        '-1' if no valid year marks are available.

        Note: a student whose *current* year has no marks (year mark = '-1') still
        gets a prior-years overall here — that is deliberate, so an MPhys/MMath Y4
        student with no Y4 result can be awarded a BSc on years 1–3.  The overall
        is only blanked to '-1' for non-awarded students (Intercal / Interrupt /
        Withdrawn etc.), done in detect_intercal / apply_special_status.

        M+P variants use the same weights as their non-M+P counterparts.
        Study abroad is flagged via is_study_abroad (auto-detected when the plan
        string contains 'study', or forced via the abroad_list override) and
        applies to both the MPhys (4) and MMath (4m) final years.
        """
        # Strip 'm' suffix so M+P variants share the same branch as non-M+P.
        base = classyear.rstrip('m')

        # Build (base_weight, mark_value) list for the years available this classyear.
        w = _BSC_WEIGHTS   # default weights for years 1–3
        if base == '1':
            candidates = [(w[1], self.yearmark)]
        elif base == '2':
            candidates = [(w[1], self.phys1), (w[2], self.yearmark)]
        elif base in ('31', '32'):
            candidates = [(w[1], self.phys1), (w[2], self.phys2), (w[3], self.yearmark)]
        elif base == '4':
            w = _MPHYS_ABROAD_WEIGHTS if self.is_study_abroad else _MPHYS_WEIGHTS
            candidates = [
                (w[1], self.phys1),
                (w[2], self.phys2),
                (w[3], self.phys3),
                (w[4], self.yearmark),
            ]
        else:
            return

        # Keep only years with a valid numeric mark (drop None and '-1' sentinel).
        valid = []
        for weight, mark in candidates:
            try:
                f = float(mark)
                if f >= 0:
                    valid.append((weight, f))
            except (TypeError, ValueError):
                pass

        if not valid:
            self.overall = '-1'
            return

        total_weight = sum(w for w, _ in valid)
        weighted_sum = sum(w * m for w, m in valid)
        self.overall = math.floor(weighted_sum / total_weight * 10 + 0.5) / 10

    def calc_project_mark(self, classyear):
        """Set project_mark (credit-weighted average, rounded int) and project_creds.

        Handles all final-year programmes:
          BSc / BSc M+P (32/32m) : BSc project (any BSC_PROJECT_MODULES code:
                                   PHYS dissertation, or MATH30022 for Maths+Physics)
          MPhys (4)              : PHYS40181 (S1) and/or PHYS40182 (S2);
                                   PHIL40000 for Physics with Philosophy students
          MMath (4m)             : MATH40011 and/or MATH40022

        If two project units exist they are combined as a credit-weighted average.
        project_mark is stored as a rounded integer (matching the 2025 convention).
        """
        base = classyear.rstrip('m')
        if base not in ('32', '4'):
            return

        def _find(name):
            """Return (numeric_mark, credits) for the first unit matching *name*."""
            for u in self.units:
                if (u.coursename or u.module or '') == name:
                    return (_numeric_mark(u.mark), u.credits)
            return (None, None)

        p1m = p1c = None
        p2m = p2c = None

        if base == '32':
            # BSc project: Physics dissertation (PHYS3088x) or MATH30022 (Maths+Physics)
            for code in BSC_PROJECT_MODULES:
                m, c = _find(code)
                if m is not None:
                    p1m, p1c = m, c
                    break

        else:
            # MPhys standard projects (S1 + S2)
            p1m, p1c = _find('PHYS40181')
            p2m, p2c = _find('PHYS40182')

            # Physics with Philosophy: PHIL40000 essay (10 cr) fills whichever slot is empty
            if p1m is None:
                p1m, p1c = _find('PHIL40000')
            elif p2m is None:
                m, c = _find('PHIL40000')
                if m is not None:
                    p2m, p2c = m, c

            # MMath: MATH40011 (S1) and/or MATH40022 (S2)
            if p1m is None:
                p1m, p1c = _find('MATH40011')
            if p2m is None:
                m, c = _find('MATH40022')
                if m is not None:
                    p2m, p2c = m, c

        # Combine into a single credit-weighted project mark
        if p1m is not None and p2m is not None:
            total_c = (p1c or 0) + (p2c or 0)
            raw = ((p1m * (p1c or 0)) + (p2m * (p2c or 0))) / total_c if total_c else (p1m + p2m) / 2
            self.project_mark  = round(raw + 0.000001)
            self.project_creds = total_c
        elif p1m is not None:
            self.project_mark  = round(p1m + 0.000001)
            self.project_creds = p1c or 0
        elif p2m is not None:
            self.project_mark  = round(p2m + 0.000001)
            self.project_creds = p2c or 0
        else:
            self.project_mark  = None
            self.project_creds = 0

    def _promo_boundary_extra(self):
        """Percent to extend each borderline-promotion zone's lower bound by.

        XB/BX in the BZ column contributes a fixed 1.0; an 'AB -n%' code beside
        any unit contributes n (the number beside it).  The two sources add
        together.  A single AB value is expected per student (repeated across
        units is fine); the largest is used, and a warning is printed only if the
        AB codes disagree on the value.
        """
        bz_codes = _split_codes(self.bz)
        extra = 1.0 if ('XB' in bz_codes or 'BX' in bz_codes) else 0.0
        ab_shifts = []
        for u in self.units:
            shift, warn = _ab_promo_shift(u.mit_circs)
            if warn:
                print(f"  WARNING: {self.emplid} {u.coursename or u.module}: {warn}")
            if shift:
                ab_shifts.append(shift)
        if len(set(ab_shifts)) > 1:
            print(f"  WARNING: {self.emplid}: conflicting 'AB' values {ab_shifts}; "
                  f"using largest ({max(ab_shifts)})")
        if ab_shifts:
            extra += max(ab_shifts)
        return extra

    def calc_deg_class(self, classyear):
        """Set deg_class_alg, deg_class_actual, borderline_for, and deg_class_rev.

        Three stages in order: (1) a mark-band class, (2) borderline promotion, then
        (3) a credit demotion applied to the *promoted* class.  This ordering matters:
        a borderline promotion that lifts a student into a higher class is still pulled
        back down if the awarded class lacks its credit requirement (e.g. promoted
        2.2 -> 2.1, then demoted 2.1 -> 2.2 on short L3 credits).
          - deg_class_alg   : the STANDARD-rules award.  A standard P(A) promotion and
                              the credit demotion are reflected (== deg_class_actual);
                              a department promotion (P(B)/P(A_X)/P(B_X)) is NOT, so Alg
                              then shows the pre-promotion class; a Y4 revert shows
                              '<programme> Fail'.
          - deg_class_actual: the final award after stages (2) and (3).
          - deg_class_rev   : 'P(A)'/'P(B)' (promotion) and/or 'CR' (a credit demotion
                              or a considered-but-not-promoted borderline), de-duplicated.
          - fail_reason     : drives the 'Award reason' column; records the credit
                              shortfall, e.g. '<80 credits (70)'.

        For BSc (classyear 32/32m):
          (1) Mark-band class from the overall boundary, requiring the MUST_PASS units
              (lab + project) and the BSC_L3_CREDITS_THIRD (60) L3+ floor:
              1/2.1/2.2/3 by mark; '3 ord' if the floor/labs are met but the mark or
              must-pass are not; 'Fail' below the L3+ floor.
          (2) Borderline zones: 2% below each upper-class boundary (BSC_BORDERLINE_UPPER),
              3% below the 3rd (BSC_BORDERLINE_THIRD); extended by _promo_boundary_extra.
              Algorithm A: >= BSC_PROMO_A_CREDITS credits (ANY level) at the target
                boundary mark -> promote, note 'P(A)'.
              Algorithm B (only if A fails): >= PROMO_B_CREDITS such credits AND the BSc
                dissertation at the target level AND yearmark > overall -> 'P(B)'.
          (3) L3+ credit demotion: the awarded class must hold BSC_L3_CREDITS_UPPER (80)
              L3+ credits; short by up to 20 (in [60, 80)) drops it ONE class
              (1->2.1->2.2->3), without cascading.

        For MPhys (classyear 4) / MMath (classyear 4m):
          (1) Mark-band class (1/2.1/2.2 — no 3rd) from the overall boundary, requiring a
              passed project; None (-> revert to BSc) below the 2.2 boundary or no project.
          (2) Borderline promotion as for BSc but with Y4_PROMO_A_CREDITS for algorithm A,
              and the MPHYS_CREDITS_SHORT credit floor.
          (3) Credit demotion: the awarded class needs MPHYS_CREDITS_FULL (200) credits
              passed (any level) over Y3+Y4; short by up to 20 (>= MPHYS_CREDITS_SHORT)
              drops one class, and a 2.2 (or anything below the short floor) reverts to a
              BSc degree based on years 1–3 (overall-mark boundary only), shown with the
              BSc Y1-Y3 mark in parentheses.  A revert is itself an MPhys/MMath fail, so
              deg_class_alg shows '<programme> Fail'; a non-reverting MPhys/MMath keeps
              its stage-(1) mark-band class.
        """
        base = classyear.rstrip('m')
        if base == '32':
            prefix = 'BSc'
        elif classyear == '4':
            prefix = 'MPhys'
        elif classyear == '4m':
            prefix = 'MMath&Phys'
        else:
            return
        orig_prefix = prefix   # programme prefix before any revert-to-BSc (MPhys/MMath fail)
        bsc_award_suffix = ''   # ' (NN.N%)' BSc Y1-Y3 mark appended to an MPhys/MMath->BSc revert

        try:
            overall = float(self.overall)
            if overall < 0:
                return
        except (TypeError, ValueError):
            return

        if base == '32':
            # Must-pass units (lab + project) not passed in this student's list;
            # has_req holds when none failed.
            must_pass_failed = {
                (u.coursename or u.module or '')
                for u in self.units
                if (u.coursename or u.module or '') in MUST_PASS
                and not _mark_accepted(u.mark)
            }
            has_req = not must_pass_failed
            l3 = self.credits_l3 + self.credits_l4   # L3+ passed credits for classification

            # --- (1) mark-band class: the algorithmic award, BEFORE promotion or the
            # L3-credit demotion — this is what 'Deg Class Alg' reports.  Honours
            # (1/2.1/2.2/3) require the must-pass units and the 60-credit L3+ floor;
            # the 80-credit requirement is applied later as a demotion (step 3), so it
            # is NOT folded into the class here.
            if has_req and l3 >= BSC_L3_CREDITS_THIRD:
                if overall >= BOUNDARY_FIRST:
                    mark_cls = '1'
                elif overall >= BOUNDARY_UPPER2:
                    mark_cls = '2.1'
                elif overall >= BOUNDARY_LOWER2:
                    mark_cls = '2.2'
                elif overall >= BOUNDARY_THIRD:
                    mark_cls = '3'
                else:
                    mark_cls = '3 ord'   # honours credits/labs met but < 40% overall
            elif l3 >= BSC_L3_CREDITS_THIRD:
                mark_cls = '3 ord'       # must-pass failed but enough L3+ credits
            else:
                mark_cls = 'Fail'

            cls     = mark_cls
            # Pre-promotion class for the Alg column: mark-band with the STANDARD L3+
            # credit demotion applied, but NO promotion.  Used only when a department
            # promotion (B/A_X/B_X) lifts the actual award; a standard P(A) promotion
            # keeps Alg == Actual.
            cls_alg = mark_cls
            if cls_alg in ('1', '2.1', '2.2') and BSC_L3_CREDITS_THIRD <= l3 < BSC_L3_CREDITS_UPPER:
                cls_alg = {'1': '2.1', '2.1': '2.2', '2.2': '3'}[cls_alg]

            # --- (2) borderline detection and promotion (UNCHANGED criteria) ---
            # Each zone maps (target class, lower bound, upper bound, credits-at-boundary attr).
            # credits_at_* count all current-year credits (any level) at the boundary mark,
            # including excluded units (computed in calc_level_credits).
            # XB/BX in the BZ column (1.0) plus any 'AB -n%' unit code (n) extend
            # every zone's lower bound (see _promo_boundary_extra).  promo_note holds the
            # review note ('P(A)'/'P(B)'/'CR'/'P(A_X)'/'P(B_X)') and is merged in step 4.
            promo_note = None
            bz_extra = self._promo_boundary_extra()
            _BORDERLINE_ZONES = (
                ('1',   BOUNDARY_FIRST  - BSC_BORDERLINE_UPPER, BOUNDARY_FIRST,  'credits_at_first'),
                ('2.1', BOUNDARY_UPPER2 - BSC_BORDERLINE_UPPER, BOUNDARY_UPPER2, 'credits_at_upper2'),
                ('2.2', BOUNDARY_LOWER2 - BSC_BORDERLINE_UPPER, BOUNDARY_LOWER2, 'credits_at_lower2'),
                ('3',   BOUNDARY_THIRD  - BSC_BORDERLINE_THIRD, BOUNDARY_THIRD,  'credits_at_third'),
            )
            for target_cls, lo, hi, at_attr in _BORDERLINE_ZONES:
                if lo - bz_extra <= overall < hi:
                    self.borderline_for = target_cls
                    creds_at_target = getattr(self, at_attr)

                    if creds_at_target >= BSC_PROMO_A_CREDITS:
                        cls = target_cls
                        promo_note = 'P(A)'
                    else:
                        # Algorithm B: project at target level AND yearmark > overall
                        try:
                            ym = float(self.yearmark)
                        except (TypeError, ValueError):
                            ym = -1.0
                        if (creds_at_target >= PROMO_B_CREDITS
                                and self.project_mark is not None and self.project_mark >= hi
                                and ym > overall):
                            cls = target_cls
                            promo_note = 'P(B)'
                        else:
                            # Record why neither algorithm promoted the student.
                            # The output column shows just 'CR'; the breakdown is kept
                            # on deg_class_rev_detail (not shown).  'CR' is always present
                            # (A's credit threshold not met).  If B's credit threshold was
                            # met, also report B's other failures.
                            reasons = ['CR']
                            if creds_at_target >= PROMO_B_CREDITS:
                                if self.project_mark is None or self.project_mark < hi:
                                    reasons.append('Proj.')
                                if ym >= 0 and not (ym > overall):
                                    reasons.append('marks')
                            promo_note = 'CR'
                            self.deg_class_rev_detail = '/'.join(reasons)

                            # --- promotion_x: informational, does not change degree class ---
                            # Targets ONE CLASS ABOVE the mark-band classification.
                            # Same rule as algorithms A and B (A tried first, then B), but the
                            # credit bar is (2/3) of assessed credits (non-excluded, with a
                            # mark; any level): A_X uses that value (capped at 80), B_X uses it
                            # minus 10 (usually 80 -> 70), keeping a 10-credit B_X fallback band.
                            _X_ONE_ABOVE = {
                                '3 ord': ('3',   'credits_at_third',  BOUNDARY_THIRD),
                                '3':     ('2.2', 'credits_at_lower2', BOUNDARY_LOWER2),
                                '2.2':   ('2.1', 'credits_at_upper2', BOUNDARY_UPPER2),
                                '2.1':   ('1',   'credits_at_first',  BOUNDARY_FIRST),
                            }
                            if mark_cls in _X_ONE_ABOVE:
                                x_cls, x_attr, x_hi = _X_ONE_ABOVE[mark_cls]
                                creds_at_target_x = getattr(self, x_attr)
                                assessed_creds = sum(
                                    u.credits for u in self.units
                                    if u.credits is not None
                                    and _numeric_mark(u.mark) is not None
                                    and not u.excluded
                                )
                                two_thirds = assessed_creds * 2 / 3
                                if creds_at_target_x >= min(two_thirds, BSC_PROMO_A_CREDITS):
                                    promo_note = 'P(A_X)'
                                elif (creds_at_target_x >= two_thirds - 10
                                        and self.project_mark is not None
                                        and self.project_mark >= x_hi
                                        and ym > overall):
                                    promo_note = 'P(B_X)'
                    break

            # --- (3) L3+ credit demotion on the (possibly promoted) class ---
            # The awarded class must hold >= BSC_L3_CREDITS_UPPER (80) L3+ credits.
            # Short by up to 20 (in [BSC_L3_CREDITS_THIRD, BSC_L3_CREDITS_UPPER)) drops
            # it ONE class.  Applied AFTER promotion, so a borderline promotion can still
            # be pulled back by a credit shortfall (e.g. promoted 2.2->2.1, then demoted
            # 2.1->2.2).  Does not cascade.  The 'Award reason' column records the shortfall.
            credit_demoted = False
            if cls in ('1', '2.1', '2.2') and BSC_L3_CREDITS_THIRD <= l3 < BSC_L3_CREDITS_UPPER:
                cls = {'1': '2.1', '2.1': '2.2', '2.2': '3'}[cls]
                credit_demoted = True
                self.fail_reason = f'<{BSC_L3_CREDITS_UPPER} credits ({l3})'

            # --- (4) review note: promotion note + 'CR' credit flag, de-duplicated ---
            rev_tokens = []
            if promo_note and promo_note != 'CR':
                rev_tokens.append(promo_note)
            if credit_demoted or promo_note == 'CR':
                rev_tokens.append('CR')
            self.deg_class_rev = ' '.join(rev_tokens) or None

            # --- fail reason for BSc Fail ---
            if cls == 'Fail':
                reasons = []
                if overall < BOUNDARY_THIRD:
                    reasons.append('< 40% overall')
                if l3 < BSC_L3_CREDITS_THIRD:
                    reasons.append(f'< {BSC_L3_CREDITS_THIRD} L3 creds ({l3})')
                if any(c in MUST_PASS_LAB for c in must_pass_failed):
                    reasons.append('Failed lab')
                if any(c in MUST_PASS_PROJECT for c in must_pass_failed):
                    reasons.append('Failed project')
                self.fail_reason = ' / '.join(reasons)
        else:
            # ----- MPhys (4) / MMath (4m) -----
            # Any-level credits passed over Y3+Y4 (incl. project), out of 240, used
            # for the MPHYS_CREDITS_FULL/SHORT thresholds.  Fall back to the
            # failed-credit subtraction if it was not pre-computed.
            credits = self.credits_passed_y3y4
            if credits is None:
                credits = MPHYS_CREDITS_TOTAL - (self.l3_l4_credits_failed or 0)
                self.credits_passed_y3y4 = credits
            project_ok = (self.project_mark is not None
                          and self.project_mark > PASS_MARK)
            if not project_ok and self.project_mark is not None:
                # Accept a project whose combined mark is below pass only because a
                # contributing unit carries a 'C'/'R' accept-suffix (compensated /
                # passed resit); every present project unit must be accepted.
                proj_units = [u for u in self.units
                              if (u.coursename or u.module or '') in PROJECT_MODULES
                              and _numeric_mark(u.mark) is not None]
                project_ok = bool(proj_units) and all(_mark_accepted(u.mark) for u in proj_units)

            # --- (1) mark-band MPhys class: algorithmic award assuming full credits,
            # BEFORE promotion or the credit demotion.  None ⇒ marks/project do not
            # support an MPhys honours ⇒ revert to BSc.  Reported in 'Deg Class Alg'
            # (shown as 'MPhys Fail' / 'MMath&Phys Fail' when None).  The credit
            # requirement is applied later as a demotion (step 3), not folded in here.
            if project_ok and overall >= BOUNDARY_FIRST:
                mark_cls = '1'
            elif project_ok and overall >= BOUNDARY_UPPER2:
                mark_cls = '2.1'
            elif project_ok and overall >= BOUNDARY_LOWER2:
                mark_cls = '2.2'
            else:
                mark_cls = None
            cls     = mark_cls
            # Pre-promotion class for the Alg column: mark-band with the STANDARD credit
            # demotion applied, but NO promotion.  Used only when a department promotion
            # (B/A_X/B_X) lifts the actual award; a standard P(A) promotion keeps
            # Alg == Actual.  A mark-band the standard rules cannot sustain on credits is
            # a fail (a revert is handled separately as '<programme> Fail').
            cls_alg = mark_cls
            if cls_alg in ('1', '2.1', '2.2'):
                if credits < MPHYS_CREDITS_SHORT:
                    cls_alg = None
                elif credits < MPHYS_CREDITS_FULL:
                    cls_alg = {'1': '2.1', '2.1': '2.2', '2.2': None}[cls_alg]
            cls_alg = cls_alg if cls_alg is not None else 'Fail'

            # --- (2) borderline promotion (UNCHANGED criteria): algorithm A, then B ---
            # A: >= Y4_PROMO_A_CREDITS credits (any level) at the target boundary mark;
            # B: >= PROMO_B_CREDITS such credits, AND the project at the target level,
            #    AND yearmark > overall.
            # Both also require the MPHYS_CREDITS_SHORT credit floor (B's project-at-target
            # test implies the project is passed).  credits_at_* are from the Y4 grid only.
            # XB/BX in the BZ column (1.0) plus any 'AB -n%' unit code (n) widen
            # every band's lower bound (see _promo_boundary_extra).
            promo_note = None
            bz_extra = self._promo_boundary_extra()
            _MPHYS_ZONES = (
                ('1',   BOUNDARY_FIRST  - BSC_BORDERLINE_UPPER, BOUNDARY_FIRST,  'credits_at_first'),
                ('2.1', BOUNDARY_UPPER2 - BSC_BORDERLINE_UPPER, BOUNDARY_UPPER2, 'credits_at_upper2'),
                ('2.2', BOUNDARY_LOWER2 - BSC_BORDERLINE_UPPER, BOUNDARY_LOWER2, 'credits_at_lower2'),
            )
            for target_cls, lo, hi, at_attr in _MPHYS_ZONES:
                if lo - bz_extra <= overall < hi:
                    self.borderline_for = target_cls
                    creds_at_target = getattr(self, at_attr)
                    try:
                        ym = float(self.yearmark)
                    except (TypeError, ValueError):
                        ym = -1.0
                    if (project_ok and credits >= MPHYS_CREDITS_SHORT
                            and creds_at_target >= Y4_PROMO_A_CREDITS):
                        cls = target_cls
                        promo_note = 'P(A)'
                    elif (credits >= MPHYS_CREDITS_SHORT
                            and creds_at_target >= PROMO_B_CREDITS
                            and self.project_mark is not None and self.project_mark >= hi
                            and ym > overall):
                        cls = target_cls
                        promo_note = 'P(B)'
                    else:
                        promo_note = 'CR'
                    break

            # --- (3) credit demotion on the (possibly promoted) class ---
            # The awarded MPhys class needs >= MPHYS_CREDITS_FULL (200) credits over
            # Y3+Y4.  Short by up to 20 (>= MPHYS_CREDITS_SHORT) drops it one class;
            # a 2.2 has no lower honours, and anything below the short floor reverts
            # to BSc (cls = None).  Applied AFTER promotion so a borderline promotion is
            # still pulled back by a credit shortfall (matching the BSc rule above).
            credit_demoted = False
            if cls in ('1', '2.1', '2.2'):
                if credits < MPHYS_CREDITS_SHORT:
                    cls = None
                    credit_demoted = True
                elif credits < MPHYS_CREDITS_FULL:
                    cls = {'1': '2.1', '2.1': '2.2', '2.2': None}[cls]
                    credit_demoted = True
                    if cls is not None:
                        self.fail_reason = f'< {MPHYS_CREDITS_FULL} Y3/Y4 creds passed ({credits})'

            if cls is None:
                # Revert to a BSc, assessed as a 3-year BSc (classyear 32): the BSc mark
                # is the BSc-weighted (Y1=10, Y2=30, Y3=60) average of the first three
                # years, and the class comes from its overall-mark boundary.  The
                # L3-credit and must-pass requirements are treated as met (the student
                # passed Y3 to progress into Y4) and borderline promotion is not
                # attempted (the Y4 grid carries no Y3 credit-at-boundary data).
                # self.overall keeps the MPhys/MMath overall (incl. Y4); the BSc Y1-Y3
                # mark is shown in parentheses after the award, e.g. 'BSc 2.2 (52.5%)'.
                prefix = 'BSc'
                bsc_candidates = [(10, self.phys1), (30, self.phys2), (60, self.phys3)]
                bsc_valid = []
                for weight, mark in bsc_candidates:
                    try:
                        f = float(mark)
                        if f >= 0:
                            bsc_valid.append((weight, f))
                    except (TypeError, ValueError):
                        pass
                if bsc_valid:
                    tw = sum(w for w, _ in bsc_valid)
                    bsc_overall = math.floor(sum(w * m for w, m in bsc_valid) / tw * 10 + 0.5) / 10
                else:
                    bsc_overall = -1.0
                bsc_award_suffix = f' ({bsc_overall:.1f}%)' if bsc_overall >= 0 else ''
                if bsc_overall >= BOUNDARY_FIRST:
                    cls = '1'
                elif bsc_overall >= BOUNDARY_UPPER2:
                    cls = '2.1'
                elif bsc_overall >= BOUNDARY_LOWER2:
                    cls = '2.2'
                elif bsc_overall >= BOUNDARY_THIRD:
                    cls = '3'
                else:
                    cls = 'Fail'

                # Explain why the MPhys/MMath was not awarded
                fail_reasons = []
                if not project_ok:
                    fail_reasons.append('Failed project')
                if overall < BOUNDARY_LOWER2:
                    fail_reasons.append(f'< {round(BOUNDARY_LOWER2)}% overall')
                else:
                    threshold = (MPHYS_CREDITS_SHORT if overall >= BOUNDARY_UPPER2
                                 else MPHYS_CREDITS_FULL)
                    if credits < threshold:
                        fail_reasons.append(
                            f'< {threshold} Y3/Y4 creds passed ({credits})'
                        )
                self.fail_reason = ' / '.join(fail_reasons)

            # --- review note: promotion note + 'CR' credit flag, de-duplicated ---
            rev_tokens = []
            if promo_note and promo_note != 'CR':
                rev_tokens.append(promo_note)
            if credit_demoted or promo_note == 'CR':
                rev_tokens.append('CR')
            self.deg_class_rev = ' '.join(rev_tokens) or None

        # 'Deg Class Actual' = final award (after promotion + demotion / revert).
        # 'Deg Class Alg' = the award the STANDARD (algorithmic) rules give:
        #   - a Y4 revert to BSc is an MPhys/MMath fail            -> '<programme> Fail'
        #   - a department (non-standard) promotion B / A_X / B_X  -> the pre-promotion
        #     class (cls_alg: mark-band with the standard credit demotion, no promotion)
        #   - everything else (a standard P(A) promotion, or no class-changing
        #     promotion, incl. a standard credit demotion)         -> same as Actual
        # The promotion/credit notes are in 'Deg Class Rev'; the credit shortfall is in
        # 'Award reason' (fail_reason).
        self.deg_class_actual = f'{prefix} {cls}{bsc_award_suffix}'
        if base == '4' and prefix == 'BSc':
            self.deg_class_alg = f'{orig_prefix} Fail'
        elif promo_note in ('P(B)', 'P(A_X)', 'P(B_X)'):
            self.deg_class_alg = f'{orig_prefix} {cls_alg}'
        else:
            self.deg_class_alg = self.deg_class_actual
        # A Y3 BSc (incl. Maths+Physics) fail is awarded an exit DipHE provided
        # the student is NOT a direct entrant into Y2 (L2CM is a real mark, not
        # the -1 'no Year-2 mark' sentinel) and is registered for a standard load
        # of 120 or 125 credits.  The algorithmic class (deg_class_alg) still
        # shows the underlying 'BSc Fail'.  A student who does not qualify for the
        # DipHE keeps 'FAIL' in the actual column.
        #
        # This is NOT applied to Y4 (base '4'): an MPhys/MMath student who fails
        # the masters reverts to a BSc on the first three years, which is the
        # floor outcome, so a Y4 fail never falls through to a DipHE.
        if base == '32' and cls == 'Fail':
            l2cm = _numeric_mark(self.phys2)
            diphe_eligible = (l2cm is not None and l2cm >= 0
                              and self.credits_taken in (120, 125))
            self.deg_class_actual = 'DipHE' if diphe_eligible else 'FAIL'

    def calc_referrals(self, classyear):
        """Determine compensation and referrals for non-final-year students.

        A lab (MUST_PASS_LAB) unit with mark == 39 is treated as a normal zone failure
        (30-39%) and flows through the same compensation/referral logic as any other unit.
        A lab unit with mark < 39 is an outright fail ('Failed lab').

        Units with EN code 'R2' were already taken as a 2nd attempt; no further
        resit can be offered.  If such a unit has mark < 30%, or is in the zone
        (30-39%), the student fails outright — 2nd-attempt units are never
        compensated.

        Y1/Y2 paths:
          Full compensation (failed <= 40 credits, no unit under 30%):
            Must-pass units → R2 (or FAIL if R2-in-EN); all others → C.
          Referral (some_unit_under_30):
            Units < 30% → R2 (or FAIL if R2-in-EN); zone must-pass/core → R2
            (or FAIL if R2-in-EN); zone non-core within 40 credits → C;
            zone non-core over cap → R2 (or FAIL if R2-in-EN).
          >40 credits, no unit under 30%:
            Must-pass/core zone units → R2 (or FAIL if R2-in-EN); non-core → C.
        """
        if classyear in FINAL_CLASSYEARS:
            return

        # Lab (MUST_PASS_LAB) units with mark == 39 (and not already at R2 attempt) are
        # zone failures, not outright fails.  R2-in-EN labs lose that exception
        # because no further resit can be offered.
        lab_near_pass_idx = set()
        r2_en_idx         = set()   # failed units whose EN column contains 'R2'
        for idx in self.failed_idx:
            unit       = self.units[idx]
            coursename = unit.coursename or unit.module
            en_codes   = _split_codes(unit.en)
            is_r2_en   = 'R2' in en_codes
            if is_r2_en:
                r2_en_idx.add(idx)
            if coursename in MUST_PASS_LAB and not is_r2_en:
                if _numeric_mark(unit.mark) == 39.0:
                    lab_near_pass_idx.add(idx)

        # A lab on exactly 39 is offered a partial (coursework-only) resit, not a
        # full resit, so its credits are NOT treated as failed for the 60-credit
        # progression threshold below.  They are still excluded from
        # credits_passed, so 'Creds Passed/Taken' is unchanged.
        lab_near_pass_creds = sum(self.units[idx].credits or 0
                                  for idx in lab_near_pass_idx)

        if classyear not in _DEFERRAL_CLASSYEARS:
            # Y3 progressing.
            if (self.credits_passed or 0) + self.credits_deferred + lab_near_pass_creds < MIN_PASS_CREDITS:
                self.fail        = True
                self.fail_reason = '< 60 credits'
                return
            # Refer lab=39 as R2; no other compensation logic yet.
            if lab_near_pass_idx:
                referred_idx     = []
                referred_courses = []
                for idx in self.failed_idx:
                    if idx in lab_near_pass_idx:
                        unit = self.units[idx]
                        coursename = unit.coursename or unit.module
                        unit.output_code = _append_code(unit.output_code, 'R2')
                        referred_idx.append(idx)
                        referred_courses.append(coursename)
                self.referred_idx     = referred_idx
                self.referred_courses = referred_courses
                self.fail_reason      = 'Resit failed lab'
                resit_parts  = [f"{c}[1]" for c in self.deferred_courses]
                resit_parts += referred_courses
                self.resits  = ' / '.join(p for p in resit_parts if p) or None
            return

        # --- Y1/Y2 FAIL checks (all collected so multiple reasons can be combined) ---
        fail_reasons = []

        if any(c in MUST_PASS_LAB
               for idx, c in zip(self.failed_idx, self.failed_courses)
               if idx not in lab_near_pass_idx):
            fail_reasons.append('Failed lab')

        if (self.credits_passed or 0) + self.credits_deferred + lab_near_pass_creds < MIN_PASS_CREDITS:
            fail_reasons.append('< 60 credits')

        # R2-in-EN units with mark < 30%: no resit available.
        for idx in r2_en_idx:
            unit = self.units[idx]
            mark = _numeric_mark(unit.mark) or 0.0
            if not (mark > MIN_MARK):
                fail_reasons.append('Failed (< 30%) 2nd attempts')
                break

        if fail_reasons:
            self.fail        = True
            self.fail_reason = ' / '.join(fail_reasons)
            return

        # All failed units flow through the same logic (lab=39 is just a zone unit).
        other_failed_idx = list(self.failed_idx)

        # --- classify all failed units ---
        failed_credits     = 0
        some_unit_under_30 = False
        for idx in other_failed_idx:
            unit            = self.units[idx]
            failed_credits += unit.credits or 0
            num = _numeric_mark(unit.mark)
            if num is None or not (num > MIN_MARK):
                some_unit_under_30 = True

        self.some_unit_under_30 = some_unit_under_30

        must_pass_for_cy = MUST_PASS_LAB | (MUST_PASS_MATHS if classyear == '1m' else frozenset())
        core_for_cy      = CORE_PHYSICS | (CORE_MATHS_PHYSICS if classyear in ('1m', '2m')
                                           else frozenset())

        zone_idx            = []
        zone_courses        = []
        compensated_idx     = []
        compensated_courses = []
        referred_idx        = []
        referred_courses    = []

        if failed_credits <= 40 and not some_unit_under_30:
            # --- full compensation path ---
            for idx in other_failed_idx:
                unit       = self.units[idx]
                coursename = unit.coursename or unit.module
                if coursename in must_pass_for_cy:
                    if idx in r2_en_idx:
                        self.fail        = True
                        self.fail_reason = 'Failed (< 30%) 2nd attempts'
                        return
                    unit.output_code = _append_code(unit.output_code, 'R2')
                    referred_idx.append(idx)
                    referred_courses.append(coursename)
                else:
                    if idx in r2_en_idx:
                        self.fail        = True
                        self.fail_reason = 'Failed 2nd attempt'
                        return
                    unit.output_code = _append_code(unit.output_code, 'C')
                    compensated_idx.append(idx)
                    compensated_courses.append(coursename)

        elif some_unit_under_30:
            # --- referral path ---
            compensation_used = 0
            for idx in other_failed_idx:
                unit       = self.units[idx]
                coursename = unit.coursename or unit.module
                mark = _numeric_mark(unit.mark) or 0.0
                if not (mark > MIN_MARK):
                    # already caught by pre-check if R2-in-EN, but defensive
                    if idx in r2_en_idx:
                        self.fail        = True
                        self.fail_reason = 'Failed (< 30%) 2nd attempts'
                        return
                    unit.output_code = _append_code(unit.output_code, 'R2')
                    referred_idx.append(idx)
                    referred_courses.append(coursename)
                else:
                    zone_idx.append(idx)
                    zone_courses.append(coursename)
                    if coursename in core_for_cy or coursename in must_pass_for_cy:
                        if idx in r2_en_idx:
                            self.fail        = True
                            self.fail_reason = 'Failed (< 30%) 2nd attempts'
                            return
                        unit.output_code = _append_code(unit.output_code, 'R2')
                        referred_idx.append(idx)
                        referred_courses.append(coursename)
                    elif compensation_used + (unit.credits or 0) <= 40:
                        if idx in r2_en_idx:
                            self.fail        = True
                            self.fail_reason = 'Failed 2nd attempt'
                            return
                        unit.output_code = _append_code(unit.output_code, 'C')
                        compensation_used += unit.credits or 0
                        compensated_idx.append(idx)
                        compensated_courses.append(coursename)
                    else:
                        if idx in r2_en_idx:
                            self.fail        = True
                            self.fail_reason = 'Failed (< 30%) 2nd attempts'
                            return
                        unit.output_code = _append_code(unit.output_code, 'R2')
                        referred_idx.append(idx)
                        referred_courses.append(coursename)

        else:
            # --- >40 credits, all in 30–39% zone (no unit below 30%) ---
            # Must-pass/core units → R2 (or FAIL if R2-in-EN); non-core → C.
            for idx in other_failed_idx:
                unit       = self.units[idx]
                coursename = unit.coursename or unit.module
                zone_idx.append(idx)
                zone_courses.append(coursename)
                if coursename in core_for_cy or coursename in must_pass_for_cy:
                    if idx in r2_en_idx:
                        self.fail        = True
                        self.fail_reason = 'Failed (< 30%) 2nd attempts'
                        return
                    unit.output_code = _append_code(unit.output_code, 'R2')
                    referred_idx.append(idx)
                    referred_courses.append(coursename)
                else:
                    if idx in r2_en_idx:
                        self.fail        = True
                        self.fail_reason = 'Failed 2nd attempt'
                        return
                    unit.output_code = _append_code(unit.output_code, 'C')
                    compensated_idx.append(idx)
                    compensated_courses.append(coursename)

        # --- Assign all results ---
        self.zone_idx            = zone_idx
        self.zone_courses        = zone_courses
        self.compensated_idx     = compensated_idx
        self.compensated_courses = compensated_courses
        self.credits_compensated = sum(self.units[i].credits or 0 for i in compensated_idx)
        self.referred_idx        = referred_idx
        self.referred_courses    = referred_courses

        if referred_idx:
            if any((self.units[i].coursename or self.units[i].module) in MUST_PASS_LAB
                   for i in referred_idx):
                self.fail_reason = 'Resit failed lab'
            resit_parts = [f"{c}[1]" for c in self.deferred_courses]
            resit_parts += referred_courses
            self.resits = ' / '.join(p for p in resit_parts if p) or None

    def __repr__(self):
        return (f'StudentInfo(emplid={self.emplid!r}, name={self.name!r}, '
                f'units={len(self.units)})')


# ===========================================================================
# Excel reading
# ===========================================================================

# Student-info columns are fixed (left-hand side of the sheet)
_STUDENT_COLS = [
    ('emplid',          0),
    ('name',            1),
    ('id_no',           2),
    ('uf',              3),
    ('mc',              4),
    ('bz',              5),
    ('admit_term',      6),
    ('entry_type',      7),
    ('psi',             8),
    ('plan',            9),
    ('units_passed',   10),
    ('award',          11),
    ('classification', 12),
]

# Sub-header names that belong to a unit block (used to find where trailing
# columns begin after the last unit)
_UNIT_SUBHEADERS = {'Module', 'Link1', 'Link2', 'Mark', 'EN', 'GBN'}


def _cell(row, col):
    """Return row[col] as a Python value, or None if missing/NaN."""
    val = row.iloc[col]
    return None if pd.isna(val) else val


def _norm_header(val):
    """Normalise a column header to a clean string key."""
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return ''
    return str(val).replace('\n', ' ').strip()


def read_students(filepath):
    """Read 'Style A Plus With Gradebook' from *filepath*.

    Returns a list of StudentInfo objects, one per student row.
    """
    df = pd.read_excel(filepath, sheet_name=SHEET_NAME,
                       header=None, dtype=object)

    unit_row   = df.iloc[UNIT_LABEL_ROW]
    header_row = df.iloc[COL_HEADER_ROW]
    n_cols     = len(header_row)

    # ------------------------------------------------------------------
    # Locate unit blocks: columns where the unit-label row has 'Unit N'
    # ------------------------------------------------------------------
    unit_starts = sorted(
        c for c, val in enumerate(unit_row)
        if isinstance(val, str) and val.startswith('Unit ')
    )
    if not unit_starts:
        raise ValueError(f"No unit columns found in {filepath}")

    # Trailing section starts at the first column after the last unit block
    # whose header is not a recognised unit sub-header.
    last_unit_start = unit_starts[-1]
    trailing_start  = n_cols   # default: no trailing
    for c in range(last_unit_start, n_cols):
        h = _norm_header(header_row.iloc[c])
        is_unit_sub = (
            h in _UNIT_SUBHEADERS
            or h.startswith('AM')
            or 'Mit' in h
            or h == ''
        )
        if not is_unit_sub:
            trailing_start = c
            break

    unit_ends = unit_starts[1:] + [trailing_start]

    # For each unit, record the absolute column index for the 4 fields
    # using fixed offsets (Module+0, Mark+3, EN+4, Mit Circs+5).
    unit_col_map = []   # list of (unit_name, module_c, mark_c, en_c, mit_c)
    for start, end in zip(unit_starts, unit_ends):
        unit_name = unit_row.iloc[start]
        cols = {field: start + offset
                for field, offset in _UNIT_COL_OFFSETS.items()
                if start + offset < end}
        unit_col_map.append((
            unit_name,
            cols.get('module'),
            cols.get('mark'),
            cols.get('en'),
            cols.get('mit_circs'),
        ))

    # Trailing column names (normalised)
    trailing_cols = [
        (c, _norm_header(header_row.iloc[c]) or f'col_{c}')
        for c in range(trailing_start, n_cols)
    ]

    # ------------------------------------------------------------------
    # Parse student rows: skip any row where emplid (col 0) is blank
    # ------------------------------------------------------------------
    # Normalised emplids for the manual study-abroad override (compared as
    # strings so the list tolerates int or string entries).
    abroad_emplids = {str(e) for e in abroad_list}

    students = []
    for _, row in df.iloc[DATA_START_ROW:].iterrows():
        if pd.isna(row.iloc[0]):
            continue

        s = StudentInfo()
        for attr, col in _STUDENT_COLS:
            setattr(s, attr, _cell(row, col))
        s.is_mphys_track  = bool(s.plan and ('MPhys' in s.plan or 'MMath' in s.plan))
        s.is_study_abroad = (
            (bool(s.plan) and 'MPhys' in s.plan and 'study' in s.plan.lower())
            or str(s.emplid) in abroad_emplids
        )

        s.units = [
            UnitInfo(
                unit_name,
                _cell(row, mc),
                _cell(row, mk),
                _cell(row, en),
                _cell(row, mit),
            )
            for unit_name, mc, mk, en, mit in unit_col_map
        ]
        # Output ordering: lab first, then project/dissertation, then the rest in
        # input order. Done here (before any per-unit index lists are built) so the
        # whole pipeline and the output grid share one consistent unit order.
        s.units = _order_units(s.units)

        s.trailing = {name: _cell(row, c) for c, name in trailing_cols}

        # 'AS Code' (Achievement Status) column → dedicated field plus flags.
        s.AS_code = s.trailing.get('AS Code')
        as_code   = str(s.AS_code).strip().upper() if s.AS_code is not None else ''
        s.RFYR    = 'RFYR' in as_code   # repeat first year        → Interrupt
        s.RYOA    = 'RYOA' in as_code   # repeat year out abroad   → Interrupt
        s.RYIA    = 'RYIA' in as_code   # repeat year in attendance → Interrupt
        s.COMP    = 'COMP' in as_code   # completed; recorded only, no special handling

        students.append(s)

    return students


# ===========================================================================
# Output configuration
# ===========================================================================

# Trailing columns written after the unit pairs, keyed by classyear.
# Values are read directly from the "2 Line format" sheet of each input file.
TRAILING_COLS = {
    '1':   ['Creds Passed/Taken', 'Year Mark', 'Status', 'Fail reason',
            'Resits', 'Notes', 'Pre-Exam Board Minutes', 'Exam Board Minutes'],
    '1m':  ['Creds Passed/Taken', 'Phys Year Mark', 'Math Year Mark', 'Year Mark',
            'Status', 'Fail reason', 'Resits', 'Notes',
            'Pre-Exam Board Minutes', 'Exam Board Minutes'],
    '2':   ['Creds Passed/Taken', 'Phys 1', 'Year Mark', 'Status', 'Fail reason',
            'Resits', 'Notes', 'Pre-Exam Board Minutes', 'Exam Board Minutes'],
    '2m':  ['Creds Passed/Taken', 'Phys 1', 'Phys Year Mark', 'Math Year Mark',
            'Year Mark', 'Status', 'Fail reason', 'Resits', 'Notes',
            'Pre-Exam Board Minutes', 'Exam Board Minutes'],
    '31':  ['Creds Passed/Taken', 'L3/L4 creds passed', 'Phys 1', 'Phys 2',
            'BZ', 'Year Mark', 'Overall', 'Status', 'Award', 'Award reason', 'Notes',
            'Pre-Exam Board Minutes', 'Exam Board Minutes'],
    '31m': ['Creds Passed/Taken', 'L3/L4 creds passed', 'Phys 1', 'Phys 2',
            'Phys Year Mark', 'Math Year Mark', 'BZ', 'Year Mark', 'Overall',
            'Status', 'Award', 'Award reason', 'Notes',
            'Pre-Exam Board Minutes', 'Exam Board Minutes'],
    '32':  ['Creds Passed/Taken', 'L3/L4 creds passed', 'Phys 1', 'Phys 2',
            'BZ', 'Year Mark', 'Overall', 'Deg Class Alg', 'Deg Class Rev',
            'Deg Class Actual', 'Award reason', 'Award', 'Classification',
            'Award Alg', 'Award Actual', 'Classification Alg',
            'Classification Actual', 'Award Change', 'Classification Change',
            'Notes', 'Pre-Exam Board Minutes', 'Exam Board Minutes'],
    '32m': ['Creds Passed/Taken', 'L3/L4 creds passed', 'Phys 1', 'Phys 2',
            'Phys Year Mark', 'Math Year Mark', 'BZ', 'Year Mark', 'Overall',
            'Deg Class Alg', 'Deg Class Rev', 'Deg Class Actual', 'Award reason',
            'Award', 'Classification', 'Award Alg', 'Award Actual',
            'Classification Alg', 'Classification Actual',
            'Award Change', 'Classification Change',
            'Notes', 'Pre-Exam Board Minutes', 'Exam Board Minutes'],
    '4':   ['Creds Passed/Taken', 'Y3 creds failed w/wo MCs',
            'L4 creds passed Y3+Y4', 'Phys 1', 'Phys 2', 'Phys 3',
            'BZ', 'Year Mark', 'Overall', 'Deg Class Alg', 'Deg Class Rev',
            'Deg Class Actual', 'Award reason', 'Award', 'Classification',
            'Award Alg', 'Award Actual', 'Classification Alg',
            'Classification Actual', 'Award Change', 'Classification Change',
            'Notes', 'Pre-Exam Board Minutes', 'Exam Board Minutes'],
    '4m':  ['Creds Passed/Taken', 'Y3 creds failed w/wo MCs',
            'L4 creds passed Y3+Y4', 'Phys 1', 'Phys 2', 'Phys 3',
            'Phys Year Mark', 'Math Year Mark', 'BZ', 'Year Mark', 'Overall',
            'Deg Class Alg', 'Deg Class Rev', 'Deg Class Actual', 'Award reason',
            'Award', 'Classification', 'Award Alg', 'Award Actual',
            'Classification Alg', 'Classification Actual',
            'Award Change', 'Classification Change',
            'Notes', 'Pre-Exam Board Minutes', 'Exam Board Minutes'],
}

# Column widths (Excel character units), measured from the "2 Line format" sheet.
# '_unit' and '_code' are the two columns of each unit pair.
_COL_WIDTHS = {
    'ID No.':                    5.00,
    'Emplid':                    9.00,
    'Name':                     12.00,
    'Plan':                     17.00,
    '_unit':                     7.00,
    '_code':                     8.00,
    'Creds Passed/Taken':       16.00,
    'Year Mark':                9.00,
    'Phys Year Mark':           12.00,
    'Math Year Mark':           12.00,
    'Status':                   11.00,
    'Fail reason':              24.00,
    'Award reason':             24.00,
    'Resits':                   50.00,
    'Notes':                    50.00,
    'Pre-Exam Board Minutes':   40.00,
    'Exam Board Minutes':       40.00,
    'Phys 1':                   7.00,
    'Phys 2':                   7.00,
    'Phys 3':                   7.00,
    'BZ':                       7.00,
    'Overall':                  8.00,
    'L3/L4 creds passed':       16.00,
    'L4 creds passed Y3+Y4':    18.00,
    'Y3 creds failed w/wo MCs': 18.00,
    'Award':                    10.00,
    'Classification':           14.00,
    'Deg Class Alg':            14.00,
    'Deg Class Rev':            14.00,
    'Deg Class Actual':         14.00,
    'Award Alg':                10.00,
    'Award Actual':             10.00,
    'Classification Alg':       16.00,
    'Classification Actual':    16.00,
    'Award Change':             10.00,
    'Classification Change':    16.00,
}

# Formatting objects (created once, reused for every cell)
_FONT        = Font(name='Aptos Narrow', size=11)
_FONT_BOLD   = Font(name='Aptos Narrow', size=11, bold=True)
_FILL_GREY   = PatternFill(fill_type='solid', fgColor='FFE0E0E0')
# Cell-highlight fills (see write_students) — medium pastels, brighter than a
# very-pale wash but softer than pure yellow (FFFF00).
_FILL_PALE_GREEN  = PatternFill(fill_type='solid', fgColor='FFA9F5A9')  # marks excluded by mitigating circumstances ('X' code)
_FILL_PALE_YELLOW = PatternFill(fill_type='solid', fgColor='FFFFF066')  # fails: Y1/Y2 in 30-39 zone, and all Y3/Y4 fails
_FILL_PALE_PINK   = PatternFill(fill_type='solid', fgColor='FFFFB3C6')  # Y1/Y2 fails below MIN_MARK (30%)
_FILL_BEIGE       = PatternFill(fill_type='solid', fgColor='FFFFD699')  # borderline yearmark / overall (light orange-tan)

# Outcome labels for students who are not assessed this cycle (already completed,
# left, intercalating, or set by hand).  Their mark cells get no status highlight
# (no green/yellow/pink/beige), and their output codes drop the R1/R2 markers.
_SPECIAL_OUTCOME_LABELS = frozenset({
    'Completed', 'Interrupt', 'Manual', 'Withdrawn', 'Intercal',
})


def _is_special_outcome(label):
    """True if *label* is a non-assessed outcome (completed/left/intercalating/
    manual), ignoring any ' (CertHE)' exit-award suffix (e.g. 'Withdrawn (CertHE)'
    counts as special, while 'FAIL (CertHE)' does not)."""
    if not label:
        return False
    return str(label).replace(' (CertHE)', '') in _SPECIAL_OUTCOME_LABELS


_ALIGN_CTR   = Alignment(horizontal='center')
_ALIGN_RIGHT = Alignment(horizontal='right')
_INFO_ROW_H  = 17.0   # height of student info rows

_SIDE_THIN    = Side(border_style='thin', color='FF808080')   # vertical column separators
_SIDE_THIN_H  = Side(border_style='thin', color='FFD0D0D0')   # horizontal row lines within a student pair
_SIDE_THIN_BK = Side(border_style='thin')                     # horizontal dividers between student pairs (black)
_SIDE_NONE    = Side()

# Fixed left-hand output columns: (header label, StudentInfo attribute name)
_FIXED_COLS = [
    ('ID No.', 'id_no'),
    ('Emplid',  'emplid'),
    ('Name',    'name'),
    ('Plan',    'plan'),
]


# Maps classyear → {StudentInfo attr: input trailing column key} for previous
# year marks.  The current year's LxCM column is intentionally excluded.
_PREV_YEARMARK_COLS = {
    '2':   {'phys1': 'L1CM'},
    '2m':  {'phys1': 'L1CM'},
    '31':  {'phys1': 'L1CM', 'phys2': 'L2CM'},
    '31m': {'phys1': 'L1CM', 'phys2': 'L2CM'},
    '32':  {'phys1': 'L1CM', 'phys2': 'L2CM'},
    '32m': {'phys1': 'L1CM', 'phys2': 'L2CM'},
    '4':   {'phys1': 'L1CM', 'phys2': 'L2CM', 'phys3': 'L3CM'},
    '4m':  {'phys1': 'L1CM', 'phys2': 'L2CM', 'phys3': 'L3CM'},
}

# Maps trailing column header names to StudentInfo attribute names.
# Populated as computed attributes are added; write_students uses this to
# fill in values rather than leaving those cells blank.
_TRAILING_ATTR = {
    'Creds Passed/Taken':    'creds_passed_taken',
    'Year Mark':             'yearmark',
    'Phys Year Mark':        'phys_yearmark',
    'Math Year Mark':        'math_yearmark',
    'Phys 1':                'phys1',
    'Phys 2':                'phys2',
    'Phys 3':                'phys3',
    'Status':                'status',
    'Fail reason':           'fail_reason',
    'Award reason':          'fail_reason',   # final years + Y3 MPhys/MMath (31/31m): same value, award-oriented header
    'Resits':                'resits',
    'BZ':                    'bz',
    'L3/L4 creds passed':       'l3_l4_creds_passed',
    'L4 creds passed Y3+Y4':    'l4_creds_y3y4_str',
    'Y3 creds failed w/wo MCs': 'y3_creds_failed_str',
    'Overall':               'overall',
    'Deg Class Alg':         'deg_class_alg',
    'Deg Class Rev':         'deg_class_rev',
    'Deg Class Actual':      'deg_class_actual',
}

# Excel number formats applied to trailing columns that hold computed floats.
_TRAILING_FORMAT = {
    'Year Mark':      '0.0',
    'Phys Year Mark': '0.0',
    'Math Year Mark': '0.0',
    'Overall':        '0.0',
    'Phys 1':         '0.0',
    'Phys 2':         '0.0',
    'Phys 3':         '0.0',
}


# ===========================================================================
# Excel writing
# ===========================================================================

def write_students(students, outpath, classyear, hide_id_cols=True):
    """Write *students* to *outpath* in 2-row-per-student format.

    Row 1    : bold header — fixed labels, 'Unit N' merged over each pair, trailing labels
    Row 2n   : student info row  — fixed fields, module codes (merged), blank trailing
    Row 2n+1 : student marks row — marks and output codes per unit, blank trailing

    Formatting:
    - Alternating grey fill (FFE0E0E0) on every other student pair
    - Medium (thick) bottom border below the header and after each student pair
    - Thin left border between every column

    Borders are applied to all cells before merging because openpyxl converts
    non-top-left merged cells to read-only MergedCell objects on merge_cells().
    """
    trailer_names = TRAILING_COLS[classyear]
    n_units  = len(students[0].units) if students else 0
    n_fixed  = len(_FIXED_COLS)
    u_start  = n_fixed + 1               # 1-based col of first unit pair
    t_start  = u_start + 2 * n_units     # 1-based col of first trailing col
    last_col = t_start - 1 + len(trailer_names)
    n_rows   = 1 + 2 * len(students)

    # Rows that get a thick bottom border: header + every marks row
    thick_rows = frozenset([1] + [2 + 2*i + 1 for i in range(len(students))])

    wb = Workbook()
    ws = wb.active
    ws.title = '2 Line Format'

    # ------------------------------------------------------------------ widths
    for i, (label, _) in enumerate(_FIXED_COLS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = _COL_WIDTHS[label]
    for i in range(n_units):
        ws.column_dimensions[get_column_letter(u_start + 2*i    )].width = _COL_WIDTHS['_unit']
        ws.column_dimensions[get_column_letter(u_start + 2*i + 1)].width = _COL_WIDTHS['_code']
    for j, name in enumerate(trailer_names):
        ws.column_dimensions[get_column_letter(t_start + j)].width = \
            _COL_WIDTHS.get(name, 12.0)

    # ------------------------------------------------------------------ header
    for i, (label, _) in enumerate(_FIXED_COLS, start=1):
        c = ws.cell(row=1, column=i, value=label)
        c.font = _FONT_BOLD
        c.alignment = _ALIGN_CTR
    for i in range(n_units):
        col = u_start + 2 * i
        c = ws.cell(row=1, column=col, value=f'Unit {i + 1}')
        c.font = _FONT_BOLD
        c.alignment = _ALIGN_CTR
        # merge deferred until after border pass
    for j, name in enumerate(trailer_names):
        c = ws.cell(row=1, column=t_start + j, value=name)
        c.font = _FONT_BOLD
        c.alignment = _ALIGN_CTR

    # ------------------------------------------------------------------ rows
    pending_merges = []   # collected here; applied after the border pass

    for idx, s in enumerate(students):
        info_row  = 2 + 2 * idx
        marks_row = info_row + 1
        fill = _FILL_GREY if idx % 2 == 1 else None

        ws.row_dimensions[info_row].height = _INFO_ROW_H

        def _c(row, col, value=None, center=False):
            cell = ws.cell(row=row, column=col, value=value)
            cell.font = _FONT
            if fill:
                cell.fill = fill
            if center:
                cell.alignment = _ALIGN_CTR
            return cell

        # info row — fixed columns (ID column bold + centered)
        for i, (_, attr) in enumerate(_FIXED_COLS, start=1):
            cell = _c(info_row, i, getattr(s, attr))
            if i == 1:
                cell.font = _FONT_BOLD
                cell.alignment = _ALIGN_CTR

        # info row — unit module codes (merge pending)
        for i, unit in enumerate(s.units):
            col = u_start + 2 * i
            _c(info_row, col,     unit.module, center=True)
            _c(info_row, col + 1, None)
            pending_merges.append((info_row, col, col + 1))

        # Students who are not assessed this cycle (completed/left/intercalating/
        # manual) get no status highlighting on any cell.
        outcome_label = s.deg_class_alg if classyear in FINAL_CLASSYEARS else s.status
        suppress_fill = _is_special_outcome(outcome_label)

        # Borderline (graduating, or progressing e.g. R/X) → beige on the mark cell
        # that drives the borderline decision.  borderline_for is set from the overall
        # mark (final years and Y3 MPhys/MMath BSc consideration, classyear 31/31m),
        # so it fills the Overall column where one exists; the Y2 'R/X' progression
        # band is yearmark-based and has no Overall column, so it fills Year Mark.
        is_borderline = (not suppress_fill
                         and (s.borderline_for is not None
                              or 'R/X' in str(s.status or '')))
        beige_tname   = 'Overall' if 'Overall' in trailer_names else 'Year Mark'

        # info row — trailing columns (computed attrs take priority; fall back to input value)
        for j, tname in enumerate(trailer_names):
            attr  = _TRAILING_ATTR.get(tname)
            value = getattr(s, attr) if attr else s.trailing.get(tname)
            if tname == 'Notes':
                value = s.cf_flags or None
            cell  = _c(info_row, t_start + j, value)
            fmt   = _TRAILING_FORMAT.get(tname)
            if fmt:
                cell.number_format = fmt
            if value == '-1':
                cell.alignment = _ALIGN_RIGHT
            if is_borderline and tname == beige_tname:
                cell.fill = _FILL_BEIGE

        # marks row — fixed columns (blank; needed for fill and borders)
        for i in range(1, n_fixed + 1):
            _c(marks_row, i)

        # marks row — unit marks and output codes
        # Fill priority: excluded by mitigating circumstances (a standalone 'X'
        # output code) → pale green.  A carried mark ('LxC') is excluded too but
        # left unfilled.  A deferral ('R1'/'XL_R1') is excluded for processing
        # but its cell is coloured by its mark like any fail.  Else a fail →
        # pale pink (Y1/Y2 below MIN_MARK) or pale yellow (Y1/Y2 in the 30-39
        # zone, and all Y3/Y4 fails).
        is_y12       = classyear in _DEFERRAL_CLASSYEARS   # Y1/Y2 classyears
        excluded_set = set(s.excluded_idx)
        failed_set   = set(s.failed_idx)
        deferred_set = set(s.deferred_idx)
        for i, unit in enumerate(s.units):
            col = u_start + 2 * i
            mark_cell = _c(marks_row, col, unit.mark)
            mark_num  = _numeric_mark(unit.mark)
            if suppress_fill:
                pass  # not assessed this cycle: no status highlight
            elif i in excluded_set:
                # Green only for mitigating-circumstances exclusions; carried
                # marks are excluded but left unfilled.  Deferrals are also held
                # in excluded_idx for processing, but their cells should still
                # show the fail colouring of the underlying mark (pink/yellow),
                # the same as a non-deferred fail — purely an output choice; the
                # deferral processing is unchanged.
                if _is_mc_excluded(unit.output_code):
                    mark_cell.fill = _FILL_PALE_GREEN
                elif (i in deferred_set and mark_num is not None
                      and mark_num <= PASS_MARK):
                    if is_y12 and mark_num < MIN_MARK:
                        mark_cell.fill = _FILL_PALE_PINK
                    else:
                        mark_cell.fill = _FILL_PALE_YELLOW
            elif i in failed_set or (mark_num is not None and mark_num <= PASS_MARK):
                if is_y12 and mark_num is not None and mark_num < MIN_MARK:
                    mark_cell.fill = _FILL_PALE_PINK
                else:
                    mark_cell.fill = _FILL_PALE_YELLOW
            _c(marks_row, col + 1, unit.output_code)

        # marks row — trailing
        for j in range(len(trailer_names)):
            _c(marks_row, t_start + j)

    # ---- border pass (must run before merge_cells) -------------------
    # thick_rows (header + every marks row) get a thin black bottom border.
    # Info rows get a lighter thin bottom border.
    # The output-code column of each unit pair has no left border so it reads
    # as one visual unit with the mark column beside it.
    code_cols = frozenset(u_start + 2*i + 1 for i in range(n_units))
    for r in range(1, n_rows + 1):
        bottom = _SIDE_THIN_BK if r in thick_rows else _SIDE_THIN_H
        for c in range(1, last_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = Border(
                left   = _SIDE_THIN if (c > 1 and c not in code_cols) else _SIDE_NONE,
                bottom = bottom,
            )

    # ---- merges (after borders) --------------------------------------
    for i in range(n_units):
        col = u_start + 2 * i
        ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + 1)
    for row, c1, c2 in pending_merges:
        ws.merge_cells(start_row=row, start_column=c1, end_row=row, end_column=c2)

    # ---- freeze first 3 columns (ID, Emplid, Name) + header row ------
    # freeze_panes pins everything above/left of the given cell; column 4,
    # row 2 keeps columns A-C visible while scrolling right and the header
    # row visible while scrolling down.
    ws.freeze_panes = f'{get_column_letter(4)}2'

    # ---- hide the 'Emplid' and 'Name' columns (unless --no_hidden) ---
    if hide_id_cols:
        for label in ('Emplid', 'Name'):
            col = next(i for i, (lbl, _) in enumerate(_FIXED_COLS, start=1)
                       if lbl == label)
            ws.column_dimensions[get_column_letter(col)].hidden = True

    wb.save(outpath)


# ===========================================================================
# Y3 credits supplementary data (for Y4 degree classification)
# ===========================================================================

def _y3cr_find_col(headers, pred, desc):
    """Return the first column index whose normalised header satisfies pred, or None."""
    for i, h in enumerate(headers):
        if pred(h.lower()):
            return i
    print(f"  WARNING: Y3 credits column '{desc}' not found")
    return None


def read_y3_credits(filepath):
    """Read supplementary Y3 credit data for Y4 students from *filepath*.

    Looks for a sheet whose name contains 'summ' (case-insensitive); falls back
    to the first sheet with a warning if none is found.  The header row is
    auto-detected as the first row (within the first 10) whose column 0 contains
    'emplid'.

    Searches column headers for:
      y3creds_below40_not_excl : contains 'not' and 'excluded', NOT 'pass'
      y3creds_below40_excl     : contains 'excluded', NOT 'not' or 'pass'
      y3creds_l4_passed        : contains 'l4' and 'pass'
      y3creds_l4_taken         : contains 'l4' and 'taken'

    Returns a dict  {emplid_str: {'y3creds_below40_not_excl': int|None,
                                  'y3creds_below40_excl':     int|None,
                                  'y3creds_l4_passed':        int|None,
                                  'y3creds_l4_taken':         int|None}}
    or an empty dict if the file is missing or unreadable.
    """
    try:
        xl = pd.ExcelFile(filepath)
    except FileNotFoundError:
        print(f"  WARNING: Y3 credits file not found — {filepath}")
        return {}
    except Exception as exc:
        print(f"  WARNING: could not read Y3 credits file {filepath} — {exc}")
        return {}

    sheet = next((s for s in xl.sheet_names if 'summ' in s.lower()), None)
    if sheet is None:
        print(f"  WARNING: no summary sheet found in {filepath}; using first sheet")
        sheet = xl.sheet_names[0]

    try:
        df = xl.parse(sheet, header=None, dtype=object)
    except Exception as exc:
        print(f"  WARNING: could not read sheet '{sheet}' from {filepath} — {exc}")
        return {}

    header_row = None
    for i in range(min(10, len(df))):
        if 'emplid' in str(df.iloc[i, 0]).lower():
            header_row = i
            break
    if header_row is None:
        print(f"  WARNING: could not find header row in sheet '{sheet}' of {filepath}")
        return {}

    headers = [_norm_header(df.iloc[header_row, c]) for c in range(len(df.columns))]

    col_not_excl = _y3cr_find_col(
        headers,
        lambda h: 'not' in h and 'excluded' in h and 'pass' not in h,
        'Y3 fails not excluded',
    )
    col_excl = _y3cr_find_col(
        headers,
        lambda h: 'excluded' in h and 'not' not in h and 'pass' not in h,
        'Y3 fails with MCs excluded',
    )
    col_l4 = _y3cr_find_col(
        headers,
        lambda h: 'l4' in h.lower() and 'pass' in h.lower(),
        'L4 passed in Y3',
    )
    col_l4_taken = _y3cr_find_col(
        headers,
        lambda h: 'l4' in h.lower() and 'taken' in h.lower(),
        'L4 taken in Y3',
    )

    def _int_val(row, col):
        if col is None:
            return None
        val = row.iloc[col]
        if pd.isna(val):
            return None
        try:
            return int(float(str(val).strip()))
        except (ValueError, TypeError):
            return None

    def _norm_emplid(v):
        try:
            return str(int(float(str(v).strip())))
        except (ValueError, TypeError):
            return str(v).strip()

    result = {}
    for _, row in df.iloc[header_row + 1:].iterrows():
        raw = row.iloc[0]
        if pd.isna(raw):
            continue
        eid = _norm_emplid(raw)
        result[eid] = {
            'y3creds_below40_not_excl': _int_val(row, col_not_excl),
            'y3creds_below40_excl':     _int_val(row, col_excl),
            'y3creds_l4_passed':        _int_val(row, col_l4),
            'y3creds_l4_taken':         _int_val(row, col_l4_taken),
        }

    return result

def read_cf_flags(filepath):
    """Read carry-forward notes from all sheets of *filepath*.

    Each sheet must have a header row and at least three columns:
      col 0 — student emplid
      col 1 — student name (ignored)
      col 2 — notes text to store in StudentInfo.cf_flags

    All sheets are read and combined; if an emplid appears on more than one
    sheet the notes are joined with '; '.

    Returns a dict {emplid_str: notes_str}, or an empty dict if the file is
    missing, disabled (None), or unreadable.
    """
    if not filepath:
        return {}
    try:
        xl = pd.ExcelFile(filepath)
    except FileNotFoundError:
        print(f"  WARNING: CF flag file not found — {filepath}")
        return {}
    except Exception as exc:
        print(f"  WARNING: could not read CF flag file {filepath} — {exc}")
        return {}

    result = {}
    for sheet in xl.sheet_names:
        try:
            df = xl.parse(sheet, header=0, dtype=object)
        except Exception:
            continue
        for _, row in df.iterrows():
            raw = row.iloc[0]
            if pd.isna(raw):
                continue
            try:
                eid = str(int(float(str(raw).strip())))
            except (ValueError, TypeError):
                eid = str(raw).strip()
            notes_raw = row.iloc[2] if len(row) > 2 else None
            if pd.isna(notes_raw):
                continue
            notes = str(notes_raw).strip()
            if eid and notes:
                result[eid] = f"{result[eid]}; {notes}" if eid in result else notes

    return result


def read_abroad_file(filepath):
    """Read study-abroad emplids from the first column of every sheet in *filepath*.

    All sheets are scanned. A cell is kept only if it looks like an emplid
    (exactly 8 digits); header rows and other non-emplid cells are skipped.
    Returns a de-duplicated list of int emplids in order of first appearance,
    or an empty list if the file is missing, disabled (None), or unreadable.
    """
    if not filepath:
        return []
    try:
        xl = pd.ExcelFile(filepath)
    except FileNotFoundError:
        print(f"  WARNING: abroad file not found — {filepath}")
        return []
    except Exception as exc:
        print(f"  WARNING: could not read abroad file {filepath} — {exc}")
        return []

    emplids = []
    seen = set()
    for sheet in xl.sheet_names:
        try:
            df = xl.parse(sheet, header=None, dtype=object)
        except Exception:
            continue
        if df.shape[1] == 0:
            continue
        for raw in df.iloc[:, 0]:
            if pd.isna(raw):
                continue
            try:
                eid = int(float(str(raw).strip()))
            except (ValueError, TypeError):
                continue   # skip header / non-numeric cells
            if len(str(eid)) == 8 and str(eid) not in seen:
                emplids.append(eid)
                seen.add(str(eid))
    return emplids


# ===========================================================================
# Argument parsing
# ===========================================================================

_REPORT_LABEL_W = 25   # label column width (padded with dots) in the report

def _lbl(text):
    """Return *text* left-padded with dots to _REPORT_LABEL_W characters."""
    return f"{text:.<{_REPORT_LABEL_W}}"


def _stats_lines(students, cy):
    """Return a list of formatted statistics lines for one processed classyear."""
    n = len(students)
    if not n:
        return ['  (no students)']

    _GRADE_ORDER  = {'1': 0, '2.1': 1, '2.2': 2, '3': 3, '3 ord': 4, 'Fail': 5}
    _PREFIX_ORDER = {'MPhys': 0, 'MMath': 0, 'MMath&Phys': 0, 'BSc': 1}
    _STATUS_ORDER = {
        'ACTV': 0, 'A/D': 1, 'REVW': 2, 'R/X': 3, 'REVW R/X': 4,
        'REVW (BSc)': 5, 'FAIL (CertHE)': 98, 'FAIL': 99,
    }

    def _deg_key(cls):
        parts = cls.split(' ', 1)
        prefix, grade = (parts[0], parts[1]) if len(parts) == 2 else ('', cls)
        return (_PREFIX_ORDER.get(prefix, 2), _GRADE_ORDER.get(grade, 9))

    def _status_key(st):
        if st.startswith('BSc'):
            # Drop a trailing ' (CR)' borderline marker before reading the grade.
            grade = st.removesuffix(' (CR)').rsplit(' ', 1)[-1]
            return (50, _GRADE_ORDER.get(grade, 9))
        return (_STATUS_ORDER.get(st, 40), 0)

    lines = []

    if cy in FINAL_CLASSYEARS:
        counts  = {}
        promo_a = promo_b = 0
        marks   = []
        for s in students:
            cls = s.deg_class_actual or '?'
            # Group MPhys->BSc reverts by class, dropping the ' (NN.N%)' BSc Y1-Y3
            # mark that the grid shows after the award (e.g. 'BSc 2.2 (52.7%)').
            cls = re.sub(r' \(\d+(?:\.\d+)?%\)$', '', cls)
            counts[cls] = counts.get(cls, 0) + 1
            rev = s.deg_class_rev or ''
            if 'P(A)' in rev and '_X' not in rev:
                promo_a += 1
            elif 'P(B)' in rev and '_X' not in rev:
                promo_b += 1
            try:
                v = float(s.overall)
                if v >= 0:
                    marks.append(v)
            except (TypeError, ValueError):
                pass

        lines.append('  Degree classification:')
        for cls in sorted(counts, key=_deg_key):
            pct = counts[cls] / n * 100
            lines.append(f"    {cls:<14}: {counts[cls]:3d}  ({pct:5.1f}%)")
        if promo_a or promo_b:
            lines.append(f"    Promoted: P(A)={promo_a}, P(B)={promo_b}")
        mark_label = 'Overall mark'
    else:
        counts = {}
        marks  = []
        for s in students:
            st = s.status or 'ACTV'
            counts[st] = counts.get(st, 0) + 1
            try:
                v = float(s.yearmark)
                if v >= 0:
                    marks.append(v)
            except (TypeError, ValueError):
                pass

        lines.append('  Status:')
        for st in sorted(counts, key=_status_key):
            pct = counts[st] / n * 100
            lines.append(f"    {st:<18}: {counts[st]:3d}  ({pct:5.1f}%)")
        mark_label = 'Year mark'

    lines.append(f"  Total students: {n}")
    if marks:
        avg = sum(marks) / len(marks)
        med = statistics.median(marks)
        lines.append(
            f"  {mark_label}: avg {avg:.1f},  median {med:.1f},"
            f"  min {min(marks):.1f},  max {max(marks):.1f}"
            + (f"  (n={len(marks)}, {n - len(marks)} without mark)"
               if len(marks) < n else f"  (n={len(marks)})")
        )
    return lines


def parse_args():
    parser = argparse.ArgumentParser(
        description=f'PyAssess2026: Physics undergraduate assessment (AY{AY})'
    )
    parser.add_argument(
        '--AY',
        type=int,
        default=None,
        metavar='YEAR',
        help=(
            f"Academic year to process (e.g. 2025, 2026), overriding the module "
            f"default (AY={AY}). Selects the matching input directory, file names "
            f"and special-status lists."
        )
    )
    parser.add_argument(
        '--classyear',
        default='all',
        metavar='CY',
        help=(
            "Class year to process: 1, 2, 31, 32, 4 "
            "(append m/M for Maths+Physics equivalent). "
            "Use 'all' or '*' to run all 10 (default)."
        )
    )
    parser.add_argument(
        '--fill_marks',
        nargs='?',
        const=50.0,
        default=FILL_MARKS,
        type=float,
        metavar='MARK',
        help=(
            "Fill blank unit marks with MARK before processing. "
            "Defaults to 50.0 when the flag is given without a value. "
            "Non-blank marks (including strings such as 'P'/'F') are left unchanged."
        )
    )
    parser.add_argument(
        '--sort', '--sort_output',
        dest='sort_output',
        action='store_true',
        default=SORT_OUTPUT,
        help=(
            "Sort output descending by mark: yearmark for progressing students, "
            "overall for final-year students. Default: %(default)s."
        )
    )
    parser.add_argument(
        '--no_hidden',
        action='store_true',
        default=False,
        help=(
            "Show the 'Emplid' and 'Name' columns in the output. "
            "By default these two columns are hidden."
        )
    )
    return parser.parse_args()


def resolve_classyears(raw):
    """Return the list of classyear keys to process."""
    val = raw.strip().lower().rstrip('/')
    if val in ('all', '*', ''):
        return ALL_CLASSYEARS
    val = val.replace('M', 'm')  # normalise upper-M
    if val not in CLASSYEAR_FILES:
        sys.exit(
            f"Error: unrecognised --classyear '{raw}'.\n"
            f"Valid values: {', '.join(ALL_CLASSYEARS)}, all, *"
        )
    return [val]


# ===========================================================================
# Main
# ===========================================================================

def apply_as_code_lists(students):
    """Augment interrupt_list / withdrawn_list from each student's AS Code.

    A student whose AS Code flags RFYR, RYOA or RYIA is added to interrupt_list
    (status → 'Interrupt'); one whose AS Code is EXIT is added to
    withdrawn_list (status → 'Withdrawn').  Emplids already on a list are not
    duplicated.  The downstream apply_special_status() then applies the label.
    Returns (n_interrupt_added, n_withdrawn_added).
    """
    seen_int = {_norm_eid(e) for e in interrupt_list}
    seen_wdr = {_norm_eid(e) for e in withdrawn_list}
    n_int = n_wdr = 0
    for s in students:
        eid  = _norm_eid(s.emplid)
        code = str(s.AS_code).strip().upper() if s.AS_code is not None else ''
        if (s.RFYR or s.RYOA or s.RYIA) and eid not in seen_int:
            interrupt_list.append(s.emplid)
            seen_int.add(eid)
            n_int += 1
        if 'EXIT' in code and eid not in seen_wdr:
            withdrawn_list.append(s.emplid)
            seen_wdr.add(eid)
            n_wdr += 1
    return n_int, n_wdr


def _missing_marks_lines(students):
    """Return warning lines listing students with genuinely missing unit marks.

    A unit counts as missing only when its mark cell is empty (see _mark_missing).
    Students on the interrupt/manual/withdrawn lists are skipped entirely.  An
    absent mark on a year-abroad unit (ABROAD_MODULES) for a study-abroad student,
    or on a placement unit (PP_MODULES) for a placement student, is expected and
    ignored; all other units are still checked.  Returns [] if no genuine gaps.
    """
    special = {_norm_eid(e) for e in (interrupt_list + manual_list + withdrawn_list)}
    lines = []
    for s in students:
        if s.COMP or _norm_eid(s.emplid) in special:
            continue
        missing = []
        for u in s.units:
            if u.module is None or not _mark_missing(u.mark):
                continue
            code = (u.coursename or '').strip()
            if (s.is_abroad and code in ABROAD_MODULES) or (s.is_pp and code in PP_MODULES):
                continue   # absent year-abroad / placement mark is expected
            missing.append(u.coursename or u.module or u.unit_name or '?')
        if missing:
            lines.append(f"    {s.emplid}: {', '.join(missing)}")
    return lines


def main():
    args = parse_args()
    if args.AY is not None and args.AY != AY:
        _configure_ay(args.AY)   # override the module default; re-derives INDIR, file maps, lists
    classyears = resolve_classyears(args.classyear)

    multi  = len(classyears) > 1
    errors = 0
    buf    = []   # collects every output line for the report file

    def _out(line=''):
        print(line)
        buf.append(line)

    # ---- header ----
    _out(f"PyAssess AY{AY}  —  processing {len(classyears)} classyear(s)")
    _out('=' * 56)

    # ---- supplementary data files ----
    y3_credits = {}
    if any(cy in ('4', '4m') for cy in classyears):
        y3cr_path = os.path.join(INDIR, Y3_CREDITS_FILE)
        y3_credits = read_y3_credits(y3cr_path)
        if y3_credits:
            _out(f"  {_lbl('Y3 credit data')}: {y3cr_path}  ({len(y3_credits)} students)")

    cf_path  = os.path.join(INDIR, CF_FLAG_FILE) if CF_FLAG_FILE else None
    cf_flags = read_cf_flags(cf_path)
    if cf_flags:
        _out(f"  {_lbl('Carry-forward notes')}: {cf_path}  ({len(cf_flags)} students)")

    # Merge any file-supplied study-abroad emplids into abroad_list (no dups);
    # read_students reads the global abroad_list, so this must run before the loop.
    abroad_path = os.path.join(INDIR, ABROAD_FILE) if ABROAD_FILE else None
    seen_abroad = {str(e) for e in abroad_list}
    added_abroad = 0
    for eid in read_abroad_file(abroad_path):
        if str(eid) not in seen_abroad:
            abroad_list.append(eid)
            seen_abroad.add(str(eid))
            added_abroad += 1
    if added_abroad:
        _out(f"  {_lbl('Study-abroad additions')}: {abroad_path}  ({added_abroad} added)")

    for cy in classyears:
        infile, outfile = CLASSYEAR_FILES[cy]
        inpath  = os.path.join(INDIR,  infile)
        outpath = os.path.join(OUTDIR, outfile)
        desc    = _CY_DESC.get(cy, cy)

        _out()
        _out(f"--- {desc}  (classyear {cy}) ---")

        try:
            students = read_students(inpath)
        except FileNotFoundError:
            _out(f"  WARNING: file not found — {inpath}")
            errors += 1
            if multi:
                continue
            else:
                sys.exit(1)
        except Exception as exc:
            _out(f"  WARNING: could not read {inpath} — {exc}")
            errors += 1
            if multi:
                continue
            else:
                sys.exit(1)

        n_units = len(students[0].units) if students else 0
        _out(f"  {_lbl('Input')}: {inpath}")
        _out(f"  {_lbl('Students')}: {len(students)} students, {n_units} units each")

        # AS Code (Achievement Status): RFYR/RYOA → Interrupt, EXIT → Withdrawn.
        n_int, n_wdr = apply_as_code_lists(students)
        if n_int or n_wdr:
            _out(f"  {_lbl('AS Code statuses')}: "
                 f"{n_int} → Interrupt, {n_wdr} → Withdrawn")

        if args.fill_marks is not None:
            filled = sum(
                1 for s in students for u in s.units if u.mark is None
            )
            for s in students:
                for u in s.units:
                    if u.mark is None:
                        u.mark = args.fill_marks
            _out(f"  {_lbl('Blank marks filled')}: {filled} with {args.fill_marks}")

        prev_cols = _PREV_YEARMARK_COLS.get(cy, {})
        for s in students:
            for attr, col_key in prev_cols.items():
                val = s.trailing.get(col_key)
                try:
                    setattr(s, attr, math.floor(float(val) * 10 + 0.5) / 10)
                except (TypeError, ValueError):
                    setattr(s, attr, '-1')
            s.exclude_units(cy)
            s.calc_yearmark(cy)
            s.calc_referrals(cy)
            s.calc_status(cy)
            if cy in _LEVEL_CREDIT_CLASSYEARS:
                s.calc_level_credits()
            s.calc_overall(cy)
            if cy in ('31', '31m'):
                s.calc_bsc_class_y3mphys(cy)
            if cy in ('4', '4m'):
                # Y3+Y4 credit accounting must run before calc_deg_class, which uses
                # credits_passed_y3y4 for MPhys/MMath classification.
                # Default to 0 for all Y4 students; override from supplementary file where matched.
                s.y3creds_below40_not_excl = 0
                s.y3creds_below40_excl     = 0
                s.y3creds_l4_passed        = 0
                s.y3creds_l4_taken         = 0
                s.y3creds_below40          = 0
                if y3_credits:
                    data = y3_credits.get(_norm_eid(s.emplid))
                    if data:
                        s.y3creds_below40_not_excl = data.get('y3creds_below40_not_excl') or 0
                        s.y3creds_below40_excl     = data.get('y3creds_below40_excl') or 0
                        s.y3creds_l4_passed        = data.get('y3creds_l4_passed') or 0
                        s.y3creds_l4_taken         = data.get('y3creds_l4_taken') or 0
                        s.y3creds_below40          = s.y3creds_below40_not_excl + s.y3creds_below40_excl

                # L3+ credits NOT yet passed in the current Y4 grid: any unit whose
                # mark is not accepted (a 'C'/'R' suffix counts as passed, see
                # _mark_accepted) — including units still awaiting a mark, which are
                # not yet earned and so must not be counted toward the credit total.
                # MC-excluded fails ARE counted, matching the Y3 side (y3creds_below40
                # sums both excluded and non-excluded Y3 fails): a unit not actually
                # passed counts against the Y3+Y4 credit total regardless of MCs.
                grid_failed = sum(
                    u.credits for u in s.units
                    if u.credits is not None
                    and not _mark_accepted(u.mark)
                    and _course_level(u.coursename or u.module or '') in (3, 4, 5, 6)
                )
                s.l3_l4_credits_failed = s.y3creds_below40 + grid_failed
                # Any-level credits passed over Y3+Y4 (incl. project), out of 240,
                # used for the MPhys/MMath award credit threshold.
                s.credits_passed_y3y4  = MPHYS_CREDITS_TOTAL - s.l3_l4_credits_failed

                # Formatted strings for output columns
                s.y3_creds_failed_str = (f"{s.y3creds_below40_excl} / "
                                         f"{s.y3creds_below40_not_excl}")
                l4_total = s.y3creds_l4_passed + s.credits_l4
                s.l4_creds_y3y4_str = f"{s.y3creds_l4_passed}+{s.credits_l4}={l4_total}"
            if cy in FINAL_CLASSYEARS:
                s.calc_project_mark(cy)
                s.calc_deg_class(cy)

            if cf_flags:
                s.cf_flags = cf_flags.get(_norm_eid(s.emplid), '')

            s.detect_intercal(cy)
            s.apply_special_status(cy)   # explicit lists override auto-detection

            # ***For testing/debugging keep this here (comment out when doing actual runs)
            #print(s.emplid, s.name)
            #if (s.emplid == 10638692):
            #    from IPython import embed
            #    embed()

        if args.sort_output:
            sort_attr = 'overall' if cy in FINAL_CLASSYEARS else 'yearmark'
            def _sort_key(s):
                label = s.deg_class_alg if cy in FINAL_CLASSYEARS else s.status
                if _is_special_outcome(label):
                    return (float('inf'), s.name or '')
                try:
                    return (-float(getattr(s, sort_attr)), s.name or '')
                except (TypeError, ValueError):
                    return (float('inf'), s.name or '')
            students.sort(key=_sort_key)

        write_students(students, outpath, cy, hide_id_cols=not args.no_hidden)
        _out(f"  {_lbl('Output')}: {outpath}")

        for line in _stats_lines(students, cy):
            _out(line)

        # Flag genuinely missing unit marks (skipped when --fill_marks is in use).
        if args.fill_marks is None:
            missing = _missing_marks_lines(students)
            if missing:
                _out(f"  WARNING: {len(missing)} student(s) with missing marks:")
                for line in missing:
                    _out(line)

    if errors and multi:
        _out(f"\n{errors} classyear(s) skipped due to missing or unreadable input files.")

    # ---- write report file ----
    if set(classyears) == set(ALL_CLASSYEARS):
        tag = 'all'
    elif len(classyears) == 1:
        tag = classyears[0]
    else:
        tag = '_'.join(classyears)
    report_path = os.path.join(INDIR, f"pyassess_results_AY{AY}_{tag}.txt")
    try:
        with open(report_path, 'w', encoding='utf-8') as fh:
            fh.write('\n'.join(buf) + '\n')
        print(f"\nReport written to: {report_path}")
    except Exception as exc:
        print(f"\nWARNING: could not write report — {exc}")

    # ***For testing/debugging keep this here (comment out when doing actual runs)
    #from IPython import embed
    #embed()

if __name__ == '__main__':
    main()
